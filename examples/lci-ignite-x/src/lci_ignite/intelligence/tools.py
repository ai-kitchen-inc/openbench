"""LCA agent tools with ContextVar render-items pattern.

Tools push A2UI visualization data to a ContextVar-backed list.
ChatEngine reads the list after agent execution via render_items_fn.
Same pattern as examples/chat/gemini_agent.py.
"""

from __future__ import annotations

import contextvars
import json
import logging
import pathlib
import re
from typing import Any

logger = logging.getLogger(__name__)

# ── Session Pipeline Store (persists across requests) ──

_session_pipelines: dict[str, dict] = {}
_current_session_var: contextvars.ContextVar[str | None] = contextvars.ContextVar(
    "lci_session_id", default=None
)


def set_pipeline_session(session_id: str) -> None:
    """Restore pipeline data for this session. Call at start of each request.

    IMPORTANT: Always creates the pipeline container in the current context
    so that asyncio.to_thread / copy_context() copies will share the same
    mutable dict. Without this, tool threads would each create their own
    container and pipeline data from parse_ldi_sheet would be invisible
    to subsequent tools like export_to_xlsx.
    """
    _current_session_var.set(session_id)
    # Always create the container in the calling context so all thread
    # copies (via ToolExecutor.execute → copy_context) share the same dict.
    container = _get_pipeline_container()
    saved = _session_pipelines.get(session_id)
    if saved is not None:
        container["data"] = saved


def _save_to_session(data: dict) -> None:
    """Persist pipeline data to the session store."""
    session_id = _current_session_var.get()
    if session_id:
        _session_pipelines[session_id] = data


# ── Per-request ContextVars ──

_render_items_var: contextvars.ContextVar[list[dict]] = contextvars.ContextVar("lci_render_items")
_current_attachments_var: contextvars.ContextVar[list[dict] | None] = contextvars.ContextVar(
    "lci_attachments", default=None
)
_upload_dir_var: contextvars.ContextVar[str] = contextvars.ContextVar(
    "lci_upload_dir", default="./uploads"
)
_pipeline_data_var: contextvars.ContextVar[dict] = contextvars.ContextVar("lci_pipeline_data")


def _get_render_list() -> list[dict]:
    """Get per-request render items list, creating if needed."""
    try:
        return _render_items_var.get()
    except LookupError:
        items: list[dict] = []
        _render_items_var.set(items)
        return items


def get_render_items() -> list[dict]:
    """Return accumulated render items from tool functions."""
    return list(_get_render_list())


def clear_render_items() -> None:
    """Clear render items queue. Called before each request."""
    _render_items_var.set([])


def set_attachments(attachments: list[dict] | None) -> None:
    """Set file attachments for the current request."""
    _current_attachments_var.set(attachments)


def set_upload_dir(upload_dir: str) -> None:
    """Set the upload directory for the current request."""
    _upload_dir_var.set(upload_dir)


def _get_pipeline_container() -> dict:
    """Get per-request pipeline data container (mutable dict).

    Uses the same pattern as _get_render_list(): a mutable container
    shared by reference across tool threads. ToolExecutor copies the
    ContextVar reference via copy_context(), so mutating the dict
    is visible to all tool calls within the same request.
    """
    try:
        return _pipeline_data_var.get()
    except LookupError:
        container: dict = {"data": None}
        _pipeline_data_var.set(container)
        return container


def clear_pipeline_data() -> None:
    """Clear pipeline data. Called before each request."""
    container = _get_pipeline_container()
    container["data"] = None


def _store_pipeline(data: dict) -> None:
    """Store processed data for the next pipeline step.

    Mutates the shared container dict (not ContextVar.set) so the
    value is visible across tool threads within the same request.
    Also persists to the session store for cross-request access.
    """
    container = _get_pipeline_container()
    container["data"] = data
    _save_to_session(data)


def _read_pipeline() -> dict | None:
    """Read data from the previous pipeline step."""
    container = _get_pipeline_container()
    return container.get("data")


def _resolve_data(data: str) -> tuple[Any, str | None]:
    """Resolve data parameter: 'auto' reads from pipeline state.

    Returns (parsed_data, error_message).
    """
    if data == "auto":
        pipeline = _read_pipeline()
        if pipeline is None:
            return None, "Error: No pipeline data available. Run parse_ldi_sheet first."
        return pipeline, None
    try:
        return json.loads(data), None
    except json.JSONDecodeError as exc:
        return None, f"Error: Invalid JSON - {exc}"


# ── File Discovery Tool ──

GET_UPLOADED_FILES_SCHEMA = {
    "type": "function",
    "function": {
        "name": "get_uploaded_files",
        "description": (
            "Get the list of files uploaded in the current request. "
            "Returns file names, disk paths, and MIME types. "
            "Call this FIRST before analyze_excel_structure or parse_ldi_sheet "
            "to get the actual file path on disk."
        ),
        "parameters": {
            "type": "object",
            "properties": {},
            "required": [],
        },
    },
}


def get_uploaded_files() -> str:
    """Return currently uploaded files with their disk paths.

    The agent must call this to discover file paths before passing
    them to analyze_excel_structure or parse_ldi_sheet.
    """
    attachments = _current_attachments_var.get()
    if not attachments:
        return json.dumps({"files": [], "message": "No files uploaded in the current request."})
    return json.dumps({"files": attachments, "count": len(attachments)})


# ── IO Table Tools ──

CREATE_IO_TABLE_SCHEMA = {
    "type": "function",
    "function": {
        "name": "create_io_table",
        "description": (
            "Create an Input-Output table from structured LCI data. "
            "Renders as ObTable in the chat interface."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "title": {
                    "type": "string",
                    "description": "Table title (e.g., 'IO Table - Cement Production')",
                },
                "process_name": {
                    "type": "string",
                    "description": "Name of the process to build the IO table for",
                },
                "data": {
                    "type": "string",
                    "description": "JSON string of process data with inputs/outputs arrays",
                },
            },
            "required": ["title", "process_name", "data"],
        },
    },
}


def create_io_table(title: str, process_name: str, data: str) -> str:
    """Create an IO table visualization from process data."""
    try:
        process_data = json.loads(data)
    except json.JSONDecodeError:
        return f"Error: Invalid JSON data for process '{process_name}'"

    inputs = process_data.get("inputs", [])
    outputs = process_data.get("outputs", [])

    headers = ["Direction", "Flow", "Category", "Amount", "Unit"]
    rows: list[list[str]] = []

    for flow in inputs:
        rows.append(
            [
                "Input",
                flow.get("flow", ""),
                flow.get("category", flow.get("section", "")),
                str(flow.get("amount", 0)),
                flow.get("unit", ""),
            ]
        )

    for flow in outputs:
        rows.append(
            [
                "Output",
                flow.get("flow", ""),
                flow.get("category", flow.get("section", "")),
                str(flow.get("amount", 0)),
                flow.get("unit", ""),
            ]
        )

    item = {"headers": headers, "rows": rows, "title": title}
    items = _get_render_list()
    items[:] = [
        i for i in items if not (i.get("title") == title and "headers" in i and "rows" in i)
    ]
    items.append(item)

    return (
        f"IO table created for '{process_name}': "
        f"{len(inputs)} inputs, {len(outputs)} outputs, {len(rows)} total rows."
    )


AGGREGATE_BY_CATEGORY_SCHEMA = {
    "type": "function",
    "function": {
        "name": "aggregate_by_category",
        "description": (
            "Aggregate flow amounts by category for a process. "
            "Useful for summarizing inputs/outputs before analysis."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "data": {
                    "type": "string",
                    "description": "JSON string of flows array [{flow, category, amount, unit}]",
                },
            },
            "required": ["data"],
        },
    },
}


def aggregate_by_category(data: str) -> str:
    """Aggregate flow amounts by category."""
    try:
        flows = json.loads(data)
    except json.JSONDecodeError:
        return "Error: Invalid JSON data"

    aggregated: dict[str, dict[str, float | str]] = {}
    for flow in flows:
        cat = flow.get("category", flow.get("section", "Unknown"))
        amount = float(flow.get("amount", 0))
        unit = flow.get("unit", "")

        if cat not in aggregated:
            aggregated[cat] = {"total": 0.0, "unit": unit, "count": 0}
        aggregated[cat]["total"] += amount
        aggregated[cat]["count"] = int(aggregated[cat]["count"]) + 1

    result = {
        cat: {"total": round(info["total"], 4), "unit": info["unit"], "count": info["count"]}
        for cat, info in aggregated.items()
    }
    return json.dumps(result, indent=2)


VALIDATE_UNITS_SCHEMA = {
    "type": "function",
    "function": {
        "name": "validate_units",
        "description": (
            "Validate unit consistency within a category. Reports any category with mixed units."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "data": {
                    "type": "string",
                    "description": "JSON string of flows array [{flow, category, amount, unit}]",
                },
            },
            "required": ["data"],
        },
    },
}


def validate_units(data: str) -> str:
    """Validate that flows in the same category use consistent units."""
    try:
        flows = json.loads(data)
    except json.JSONDecodeError:
        return "Error: Invalid JSON data"

    category_units: dict[str, set[str]] = {}
    for flow in flows:
        cat = flow.get("category", flow.get("section", "Unknown"))
        unit = flow.get("unit", "")
        if cat not in category_units:
            category_units[cat] = set()
        category_units[cat].add(unit)

    issues = []
    for cat, units in category_units.items():
        if len(units) > 1:
            issues.append(f"Category '{cat}' has mixed units: {sorted(units)}")

    if issues:
        return "Unit validation issues found:\n" + "\n".join(issues)
    return "All categories have consistent units."


CREATE_IO_TABLE_CHART_SCHEMA = {
    "type": "function",
    "function": {
        "name": "create_io_table_chart",
        "description": (
            "Create a bar chart visualization of IO table data, "
            "showing flow amounts grouped by category."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "title": {"type": "string", "description": "Chart title"},
                "data": {
                    "type": "string",
                    "description": (
                        "JSON string of chart data array [{category, amount, direction}]"
                    ),
                },
            },
            "required": ["title", "data"],
        },
    },
}


def create_io_table_chart(title: str, data: str) -> str:
    """Create a bar chart for IO table data."""
    try:
        chart_data = json.loads(data)
    except json.JSONDecodeError:
        return "Error: Invalid JSON data for chart"

    item: dict[str, Any] = {
        "type": "bar",
        "title": title,
        "data": chart_data,
    }

    items = _get_render_list()
    items[:] = [i for i in items if not (i.get("title") == title and "data" in i)]
    items.append(item)

    return f"IO chart created: '{title}' with {len(chart_data)} data points."


# ── Hotspot Analysis Tools ──

CALCULATE_PARETO_SCHEMA = {
    "type": "function",
    "function": {
        "name": "calculate_pareto",
        "description": (
            "Perform Pareto analysis on environmental impact data. "
            "Identifies the top contributors that account for the given "
            "threshold percentage (default 80%) of total impact."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "impacts": {
                    "type": "string",
                    "description": (
                        "JSON string of impact data array [{name, amount, unit, category}]"
                    ),
                },
                "threshold": {
                    "type": "number",
                    "description": "Cumulative percentage threshold (default 80.0)",
                },
            },
            "required": ["impacts"],
        },
    },
}


def calculate_pareto(impacts: str, threshold: float = 80.0) -> str:
    """Perform Pareto analysis on environmental impact data.

    Sorts impacts by amount descending, calculates cumulative percentage,
    and identifies items contributing to the threshold percentage.
    """
    try:
        impact_list = json.loads(impacts)
    except json.JSONDecodeError:
        return "Error: Invalid JSON data for Pareto analysis"

    if not impact_list:
        return "Error: Empty impact data"

    # Sort by absolute amount descending
    sorted_impacts = sorted(impact_list, key=lambda x: abs(float(x.get("amount", 0))), reverse=True)

    total = sum(abs(float(item.get("amount", 0))) for item in sorted_impacts)
    if total == 0:
        return "Error: Total impact is zero"

    cumulative = 0.0
    hotspots = []
    non_hotspots = []

    for item in sorted_impacts:
        amount = abs(float(item.get("amount", 0)))
        pct = (amount / total) * 100
        cumulative += pct

        entry = {
            "name": item.get("name", item.get("flow", "")),
            "amount": float(item.get("amount", 0)),
            "unit": item.get("unit", ""),
            "category": item.get("category", ""),
            "percentage": round(pct, 2),
            "cumulative_percentage": round(cumulative, 2),
            "is_hotspot": cumulative <= threshold or len(hotspots) == 0,
        }

        if entry["is_hotspot"]:
            hotspots.append(entry)
        else:
            non_hotspots.append(entry)

    result = {
        "hotspots": hotspots,
        "non_hotspots": non_hotspots,
        "total": round(total, 4),
        "threshold": threshold,
        "hotspot_count": len(hotspots),
        "hotspot_percentage": round(sum(h["percentage"] for h in hotspots), 2),
    }

    return json.dumps(result, indent=2)


CREATE_PARETO_CHART_SCHEMA = {
    "type": "function",
    "function": {
        "name": "create_pareto_chart",
        "description": (
            "Create a Pareto chart showing environmental impact distribution. "
            "Bar chart with cumulative percentage line."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "title": {"type": "string", "description": "Chart title"},
                "data": {
                    "type": "string",
                    "description": (
                        "JSON string of Pareto data array "
                        "[{name, percentage, cumulative_percentage}]"
                    ),
                },
            },
            "required": ["title", "data"],
        },
    },
}


def create_pareto_chart(title: str, data: str) -> str:
    """Create a Pareto chart visualization."""
    try:
        chart_data = json.loads(data)
    except json.JSONDecodeError:
        return "Error: Invalid JSON data for Pareto chart"

    item: dict[str, Any] = {
        "type": "bar",
        "title": title,
        "data": chart_data,
        "options": {
            "xKey": "name",
            "series": [
                {"dataKey": "percentage", "name": "Impact %"},
            ],
        },
    }

    items = _get_render_list()
    items[:] = [i for i in items if not (i.get("title") == title and "data" in i)]
    items.append(item)

    return f"Pareto chart created: '{title}' with {len(chart_data)} items."


CREATE_HOTSPOT_TABLE_SCHEMA = {
    "type": "function",
    "function": {
        "name": "create_hotspot_table",
        "description": (
            "Create a table summarizing environmental hotspots. "
            "Renders as ObTable in the chat interface."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "title": {"type": "string", "description": "Table title"},
                "headers": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Column headers",
                },
                "rows": {
                    "type": "array",
                    "items": {"type": "array", "items": {"type": "string"}},
                    "description": "Table rows (2D array of strings)",
                },
            },
            "required": ["title", "headers", "rows"],
        },
    },
}


def create_hotspot_table(title: str, headers: list[str], rows: list[list[str]]) -> str:
    """Create a hotspot summary table."""
    item = {"headers": headers, "rows": rows, "title": title}
    items = _get_render_list()
    items[:] = [
        i for i in items if not (i.get("title") == title and "headers" in i and "rows" in i)
    ]
    items.append(item)

    return f"Hotspot table created: '{title}' with {len(headers)} columns and {len(rows)} rows."


CREATE_HOTSPOT_CALLOUT_SCHEMA = {
    "type": "function",
    "function": {
        "name": "create_hotspot_callout",
        "description": (
            "Display a callout box highlighting key findings from hotspot analysis. "
            "Renders as ObCallout in the chat interface."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "content": {
                    "type": "string",
                    "description": "Callout content (markdown supported)",
                },
                "variant": {
                    "type": "string",
                    "enum": ["default", "info", "warning", "error", "success"],
                    "description": "Callout style variant (default: 'warning')",
                },
                "title": {
                    "type": "string",
                    "description": "Optional callout title",
                },
            },
            "required": ["content"],
        },
    },
}


def create_hotspot_callout(content: str, variant: str = "warning", title: str = "") -> str:
    """Display a callout highlighting hotspot findings."""
    item: dict[str, Any] = {"calloutContent": content, "variant": variant}
    if title:
        item["title"] = title
    items = _get_render_list()
    items[:] = [i for i in items if "calloutContent" not in i]
    items.append(item)

    return f"Hotspot callout displayed: '{title or variant}' variant."


# ── Narrative Tools ──

CREATE_NARRATIVE_MARKDOWN_SCHEMA = {
    "type": "function",
    "function": {
        "name": "create_narrative_markdown",
        "description": (
            "Render a narrative explanation section as rich markdown. "
            "Used for hotspot explanations and PROPER 2025 recommendations."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "title": {"type": "string", "description": "Section title"},
                "content": {
                    "type": "string",
                    "description": "Markdown content for the narrative section",
                },
            },
            "required": ["title", "content"],
        },
    },
}


def create_narrative_markdown(title: str, content: str) -> str:
    """Render a narrative markdown section. No A2UI push — text goes through streaming."""
    return f"## {title}\n\n{content}"


CREATE_NARRATIVE_CALLOUT_SCHEMA = {
    "type": "function",
    "function": {
        "name": "create_narrative_callout",
        "description": (
            "Display a callout for key narrative insights or PROPER 2025 recommendations."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "content": {"type": "string", "description": "Callout content"},
                "variant": {
                    "type": "string",
                    "enum": ["default", "info", "warning", "success"],
                    "description": "Callout style (default: 'info')",
                },
                "title": {"type": "string", "description": "Optional title"},
            },
            "required": ["content"],
        },
    },
}


def create_narrative_callout(content: str, variant: str = "info", title: str = "") -> str:
    """Display a callout for narrative insights."""
    item: dict[str, Any] = {"calloutContent": content, "variant": variant}
    if title:
        item["title"] = title
    items = _get_render_list()
    # Allow multiple callouts in narrative (unlike hotspot which replaces)
    items.append(item)

    return f"Narrative callout displayed: '{title or variant}'."


# ── Export Tool ──

EXPORT_TO_DOCX_SCHEMA = {
    "type": "function",
    "function": {
        "name": "export_to_docx",
        "description": (
            "Export the LCA analysis results to a .docx report file. "
            "Triggers DocxReportGenerator and shows a file download card."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "title": {"type": "string", "description": "Report title"},
                "content": {
                    "type": "string",
                    "description": "JSON string of report content sections",
                },
                "filename": {
                    "type": "string",
                    "description": "Output filename (default: lca_report.docx)",
                },
            },
            "required": ["title", "content"],
        },
    },
}


def export_to_docx(title: str, content: str, filename: str = "lca_report.docx") -> str:
    """Export LCA report to .docx format using DocxReportGenerator.

    Generates an actual .docx file in the uploads directory and pushes
    a file download card to the render items list.
    """
    from lci_ignite.output.docx_generator import DocxReportGenerator

    upload_dir = _upload_dir_var.get()
    output_path = str(pathlib.Path(upload_dir) / filename)

    # Parse content JSON
    try:
        sections = json.loads(content)
    except (json.JSONDecodeError, TypeError):
        sections = {"narrative": content}

    # Generate the actual .docx file
    generator = DocxReportGenerator()
    try:
        result = generator.generate(
            content=sections,
            output_path=output_path,
            title=title,
        )
        file_size = result.size_bytes
        logger.info("Generated DOCX report: %s (%d bytes)", output_path, file_size)
    except Exception as e:
        logger.error("Failed to generate DOCX: %s", e)
        return f"Error generating report: {e}"

    # Push file card to render items
    item: dict[str, Any] = {
        "name": filename,
        "url": f"/uploads/{filename}",
        "mimeType": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "size": file_size,
    }
    items = _get_render_list()
    items[:] = [
        i for i in items if not (i.get("name") == filename and "url" in i and "mimeType" in i)
    ]
    items.append(item)

    return f"Report generated: '{title}' → {filename} ({file_size:,} bytes)"


# ── Excel Export Tool ──

EXPORT_TO_XLSX_SCHEMA = {
    "type": "function",
    "function": {
        "name": "export_to_xlsx",
        "description": (
            "Export the IO Table to an Excel (.xlsx) file in PROPER format. "
            "Uses pipeline data automatically. Shows a file download card."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "data": {
                    "type": "string",
                    "description": "Use 'auto' to chain from pipeline (default)",
                    "default": "auto",
                },
                "filename": {
                    "type": "string",
                    "description": "Output filename (default: io_table.xlsx)",
                    "default": "io_table.xlsx",
                },
                "title": {
                    "type": "string",
                    "description": "IO Table title",
                    "default": "IO Table PROPER",
                },
            },
            "required": [],
        },
    },
}


def export_to_xlsx(
    data: str = "auto", filename: str = "io_table.xlsx", title: str = "IO Table PROPER"
) -> str:
    """Export IO Table to Excel (.xlsx) matching PROPER template layout.

    Layout (1-indexed columns):
        A: empty
        B: Input/Output (flow name or section header)
        C: Total (amount, sesuai periode kajian)
        D: Unit
        E+ per product: Jumlah/FU | Unit | %
    Flows are already aggregated by material — no Area/Process column.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from lci_ignite.data.lci_schema import IO_TABLE_SECTION_ORDER

    parsed, err = _resolve_data(data)
    if err:
        return err
    flows = _extract_flows(parsed)

    # Determine products and FU config from pipeline
    pipeline = _read_pipeline()
    products: list[dict] = []
    fu_unit_labels: dict[str, str] = {}
    fu_mode = "per_mj"
    if pipeline:
        prod_list = pipeline.get("products", [])
        if prod_list:
            if isinstance(prod_list[0], str):
                products = [{"name": n} for n in prod_list]
            elif isinstance(prod_list[0], dict):
                products = prod_list
        fu_unit_labels = pipeline.get("fu_unit_labels", {})
        fu_mode = pipeline.get("fu_mode", "per_mj")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IO Table"

    # ── Column offsets (1-indexed) ──
    COL_B = 2  # Input/Output
    COL_C = 3  # Total
    COL_D = 4  # Unit
    COL_PROD_START = 5  # First product column (E)
    COLS_PER_PRODUCT = 3  # Jumlah/FU, Unit, %
    total_cols = COL_D + len(products) * COLS_PER_PRODUCT  # last used column

    # ── Styles ──
    header_font = Font(bold=True, size=11)
    section_font = Font(bold=True, size=10)
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_font = Font(bold=True, italic=True, size=10)
    num_align = Alignment(horizontal="right")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Row 1: empty ──
    # (intentionally blank)

    # ── Row 2: product names at their column positions ──
    for i, product in enumerate(products):
        name = product.get("name", "Product")
        col = COL_PROD_START + i * COLS_PER_PRODUCT
        ws.cell(row=2, column=col, value=name).font = header_font

    # ── Row 3: "ALL PHM" label + reference value per product ──
    ws.cell(row=3, column=COL_D, value="ALL PHM").font = header_font
    for i, product in enumerate(products):
        if fu_mode == "per_output_unit":
            fu_factor = product.get("fu_unit_factor", 0)
            energy = product.get("total_energy_mj", 0)
            ref_val = energy / fu_factor if fu_factor else energy
        else:
            ref_val = product.get("total_energy_mj", "")
        col = COL_PROD_START + i * COLS_PER_PRODUCT
        if ref_val:
            ws.cell(row=3, column=col, value=ref_val).alignment = num_align

    # ── Row 4: main header row ──
    ws.cell(row=4, column=COL_B, value="Input/Output").font = header_font
    ws.cell(row=4, column=COL_B).alignment = center_align
    ws.cell(row=4, column=COL_C, value="Total").font = header_font
    ws.cell(row=4, column=COL_C).alignment = center_align
    ws.cell(row=4, column=COL_D, value="Unit").font = header_font
    ws.cell(row=4, column=COL_D).alignment = center_align
    # Product name headers merged across 3 columns
    for i, product in enumerate(products):
        name = product.get("name", "Product")
        col = COL_PROD_START + i * COLS_PER_PRODUCT
        ws.cell(row=4, column=col, value=name).font = header_font
        ws.cell(row=4, column=col).alignment = center_align
        ws.merge_cells(
            start_row=4,
            start_column=col,
            end_row=4,
            end_column=col + COLS_PER_PRODUCT - 1,
        )
    # Merge vertically: B4:B5, D4:D5
    ws.merge_cells(start_row=4, start_column=COL_B, end_row=5, end_column=COL_B)
    ws.merge_cells(start_row=4, start_column=COL_D, end_row=5, end_column=COL_D)

    # ── Row 5: sub-headers ──
    ws.cell(row=5, column=COL_C, value="(sesuai periode kajian)").font = Font(size=9)
    for i, _ in enumerate(products):
        col = COL_PROD_START + i * COLS_PER_PRODUCT
        ws.cell(row=5, column=col, value="Jumlah/FU").font = header_font
        ws.cell(row=5, column=col + 1, value="Unit").font = header_font
        ws.cell(row=5, column=col + 2, value="%").font = header_font

    # ── Group flows by category ──
    from collections import defaultdict

    by_category: dict[str, list[dict]] = defaultdict(list)
    for flow in flows:
        cat = flow.get("category", "Unknown")
        by_category[cat].append(flow)

    # Build emission sub-sections from Emisi Udara flows
    emisi_flows = by_category.pop("Emisi Udara", [])
    if emisi_flows:
        emission_sections = _build_emission_subsections(emisi_flows, products)
        for sec_name, sec_flows in emission_sections.items():
            by_category[sec_name] = sec_flows

    # Track which sections are emission summary (no % or process)
    emission_summary_section = "Emisi Udara"

    row_idx = 6  # data starts at row 6

    def _write_xlsx_section(section: str, section_flows: list[dict]) -> None:
        nonlocal row_idx
        if not section_flows:
            return

        is_emission_summary = section == emission_summary_section

        # Section header row
        cell = ws.cell(row=row_idx, column=COL_B, value=section)
        cell.font = section_font
        cell.fill = section_fill
        for c in range(COL_B, total_cols + 1):
            ws.cell(row=row_idx, column=c).fill = section_fill
        row_idx += 1

        sorted_flows = sorted(section_flows, key=lambda f: abs(f.get("amount", 0)), reverse=True)

        section_total = 0.0
        for flow in sorted_flows:
            amount = flow.get("amount", 0.0)
            section_total += amount
            ws.cell(row=row_idx, column=COL_B, value=flow.get("flow_name", ""))
            ws.cell(row=row_idx, column=COL_C, value=amount).alignment = num_align
            ws.cell(row=row_idx, column=COL_D, value=flow.get("unit", ""))

            for i, product in enumerate(products):
                name = product.get("name", "")
                fu_key = f"fu_per_mj_{name}"
                pct_key = f"pct_{name}"
                fu_val = flow.get(fu_key, 0.0)
                pct_val = flow.get(pct_key, 0.0)
                fu_label = fu_unit_labels.get(name, "MJ")
                fu_unit = f"{flow.get('unit', '')}/{fu_label}" if flow.get("unit") else ""

                col = COL_PROD_START + i * COLS_PER_PRODUCT
                if fu_val:
                    ws.cell(row=row_idx, column=col, value=fu_val).alignment = num_align
                ws.cell(row=row_idx, column=col + 1, value=fu_unit)
                if not is_emission_summary and pct_val:
                    ws.cell(row=row_idx, column=col + 2, value=pct_val).alignment = num_align

            row_idx += 1

        # Total row
        if len(sorted_flows) > 1:
            cell = ws.cell(row=row_idx, column=COL_B, value="Total")
            cell.font = total_font
            ws.cell(row=row_idx, column=COL_C, value=section_total).font = total_font
            ws.cell(row=row_idx, column=COL_C).alignment = num_align
            total_unit = sorted_flows[0].get("unit", "")
            ws.cell(row=row_idx, column=COL_D, value=total_unit).font = total_font

            for i, product in enumerate(products):
                name = product.get("name", "")
                fu_key = f"fu_per_mj_{name}"
                fu_total = sum(f.get(fu_key, 0.0) for f in sorted_flows)
                fu_label = fu_unit_labels.get(name, "MJ")
                fu_unit = f"{sorted_flows[0].get('unit', '')}/{fu_label}"
                col = COL_PROD_START + i * COLS_PER_PRODUCT
                ws.cell(row=row_idx, column=col, value=fu_total).font = total_font
                ws.cell(row=row_idx, column=col).alignment = num_align
                ws.cell(row=row_idx, column=col + 1, value=fu_unit).font = total_font
                if not is_emission_summary:
                    ws.cell(row=row_idx, column=col + 2, value=1).font = total_font
            row_idx += 1

    for section in IO_TABLE_SECTION_ORDER:
        _write_xlsx_section(section, by_category.get(section, []))

    # Auto-width columns (skip column A which is intentionally empty)
    for col_idx in range(COL_B, total_cols + 1):
        max_len = 10
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_row=1, max_row=row_idx, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
    # Column A narrow
    ws.column_dimensions["A"].width = 2

    # Save
    upload_dir = _upload_dir_var.get()
    output_path = str(pathlib.Path(upload_dir) / filename)
    wb.save(output_path)
    file_size = pathlib.Path(output_path).stat().st_size

    # Push file card to render items
    item: dict[str, Any] = {
        "name": filename,
        "url": f"/uploads/{filename}",
        "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "size": file_size,
    }
    items = _get_render_list()
    items[:] = [
        i for i in items if not (i.get("name") == filename and "url" in i and "mimeType" in i)
    ]
    items.append(item)

    return (
        f"Excel IO Table exported: '{title}' -> {filename} "
        f"({row_idx - 1} rows, {total_cols} columns, {file_size:,} bytes)"
    )


# ── LLM Auto-Mapping Tool (Layer 3) ──

# ContextVar to pass the LLM model name to the tool
_llm_model_var: contextvars.ContextVar[str] = contextvars.ContextVar(
    "lci_llm_model", default="gemini-2.5-flash"
)


def set_llm_model(model: str) -> None:
    """Set the LLM model name for generate_mapping_profile."""
    _llm_model_var.set(model)


_MAPPING_PROFILE_PROMPT = """\
You are an LCA data expert. Given the Excel file structure below, generate a MappingProfile JSON \
that maps the columns to the Standard LCI Schema.

## Excel File Structure

{excel_profile_json}

## Standard LCI Categories (Indonesian)

INPUTS: Bahan Baku, Air, Bahan Pendukung Cairan, Bahan Pendukung Padatan, \
Transportasi Bahan Bakar dan Bahan Pendukung, Fuel Gas, Bahan Bakar Cair, Listrik, \
Infrastruktur, Lahan Digunakan, Lahan Ditransformasi

OUTPUTS: Produk, Sampah, Limbah B3, Limbah Cair, Kandungan Limbah Cair, Emisi Udara

EXCLUDED (skip these): Raw Material from Processes, Other Supporting Material
HELPER (extract as helper_data): Projected Lifetime of Infrastructure, Projected Lifetime of Land

## English ↔ Indonesian Category Mapping

Raw Material from Nature → Bahan Baku
Water → Air
Liquid Supporting Material → Bahan Pendukung Cairan
Solid Supporting Material → Bahan Pendukung Padatan
Transport of Supporting Material → Transportasi Bahan Bakar dan Bahan Pendukung
Transportation of Supporting Material → Transportasi Bahan Bakar dan Bahan Pendukung
Fuel Gas → Fuel Gas
Liquid Fuels → Bahan Bakar Cair
Electricity → Listrik
Infrastructure → Infrastruktur
Land / Land Used → Lahan Digunakan
Product → Produk
Co-Product → Co-Product
Non-Hazardous Waste → Limbah Non-B3
Hazardous Waste → Limbah B3
Liquid Waste → Limbah Cair
Liquid Waste Substances → Kandungan Limbah Cair
Air Emissions → Emisi Udara

## Unit Conversions (standard)

- ton → kg (factor: 1000) for: Bahan Pendukung Padatan,
  Limbah Non-B3, Limbah B3, Emisi Udara, Infrastruktur
- barrel → L (factor: 158.987) for: Air
- m3 → L (factor: 1000) for: Air

## Instructions

Analyze the headers and sample data to produce a MappingProfile JSON with these fields:

1. **profile_name**: A slug derived from company/sheet name (lowercase, underscores)
2. **company**: Company name (from sheet name or data)
3. **scope**: Scope/zone from sheet name or data
4. **sheet_name**: Exact sheet name to parse
5. **expected_headers**: First ~19 common headers from the header row
6. **column_mapping**: Map each semantic field to its 0-based column index:
   - process, category, flow_name, direction, unit (REQUIRED)
   - scope_value (the main amount column, usually named after the company/scope)
   - total_bulk (if present)
   - For each product: per_product_<name> and fu_<name> columns
7. **header_row**: 1-based row number of the header row
8. **products**: Array of product objects detected from
   "Total per Product *" and "Functional Unit *" columns:
   - name: Product name (from column header)
   - column: Key in column_mapping for per-product amount
   - fu_column: Key in column_mapping for functional unit value
   - total_energy_mj: 0 (unknown, user must provide later)
   - fu_unit_factor: 0 (unknown, user must provide later)
   - output_unit: Product unit (from Product rows in data)
9. **category_mapping**: Map English LDI categories found in data → Indonesian standard names \
(use the mapping above, set null for excluded/helper categories)
10. **unit_conversions**: Standard conversions (see above), only include categories present in data
11. **study_period**: {{"years": 1, "description": "Annual study period"}}

## Rules

- Look at the "LDI Category" or similar column in sample data to find category names used
- The amount column is usually named after the company scope (e.g., "Pusri IB", "Semberah EP")
- "Total per Product *" columns contain per-product amounts
- "Functional Unit *" columns contain per-MJ functional unit values
- Some files have a "No" column at index 0, some don't — check the headers
- Set total_energy_mj and fu_unit_factor to 0 for products (unknown without external data)
- Only include unit_conversions for categories actually present in the data

Return ONLY valid JSON, no markdown fences, no explanations.
"""


def _call_llm_for_profile(prompt: str) -> tuple[str, str | None]:
    """Call Gemini LLM to generate a MappingProfile.

    Returns (response_text, error). If error is not None, response_text is empty.
    Extracted as a function for easy mocking in tests.
    """
    import os

    try:
        from google import genai
        from google.genai import types
    except ImportError:
        return "", "google-genai package not installed. Install with: pip install google-genai"

    api_key = os.environ.get("GOOGLE_API_KEY", "")
    if not api_key:
        return "", "GOOGLE_API_KEY not set. Cannot generate profile."

    try:
        client = genai.Client(api_key=api_key)
        model = _llm_model_var.get()

        response = client.models.generate_content(
            model=model,
            contents=prompt,
            config=types.GenerateContentConfig(
                response_mime_type="application/json",
                temperature=0.1,
            ),
        )
        return response.text.strip(), None
    except Exception as exc:
        return "", f"LLM call failed: {exc}"


GENERATE_MAPPING_PROFILE_SCHEMA = {
    "type": "function",
    "function": {
        "name": "generate_mapping_profile",
        "description": (
            "Generate a MappingProfile for an unknown Excel LDI file using LLM. "
            "Reads the file structure (headers, sample rows) and produces a column mapping. "
            "The profile is saved for reuse with future uploads of the same format. "
            "Call this when analyze_excel_structure shows no matching profile."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Path to the Excel file (.xlsx)",
                },
            },
            "required": ["file_path"],
        },
    },
}


def generate_mapping_profile(file_path: str) -> str:
    """Generate a MappingProfile using LLM from Excel structure (Layer 3).

    1. Extract ExcelProfile (Layer 2)
    2. Call LLM with structured prompt
    3. Validate and save the generated profile
    4. Return profile name for use with parse_ldi_sheet
    """
    from lci_ignite.data.excel_profile import ExcelProfile
    from lci_ignite.data.mapping_profiles import save_profile

    # Step 1: Extract Excel structure
    try:
        excel_profile = ExcelProfile.extract(file_path)
    except (FileNotFoundError, ValueError) as exc:
        return json.dumps({"error": f"Failed to read Excel file: {exc}"})

    # Step 2: Build prompt with Excel profile data
    profile_json = json.dumps(excel_profile, indent=2, default=str)
    prompt = _MAPPING_PROFILE_PROMPT.format(excel_profile_json=profile_json)

    # Step 3: Call LLM
    response_text, error = _call_llm_for_profile(prompt)
    if error:
        return json.dumps({"error": error})

    # Step 4: Parse and validate the generated profile
    try:
        generated_profile = json.loads(response_text)
    except json.JSONDecodeError as exc:
        return json.dumps(
            {
                "error": f"LLM returned invalid JSON: {exc}",
                "raw_response": response_text[:500],
            }
        )

    # Normalize column_mapping: LLM may return {key: int} instead of {key: {index: int}}
    generated_profile = _normalize_generated_profile(generated_profile)

    # Basic validation
    validation_errors = _validate_generated_profile(generated_profile)
    if validation_errors:
        return json.dumps(
            {
                "error": "Generated profile has validation errors",
                "validation_errors": validation_errors,
                "generated_profile": generated_profile,
            }
        )

    # Step 5: Sanitize name and save profile
    from lci_ignite.data.mapping_profiles import _sanitize_profile_name

    raw_name = generated_profile.get("profile_name", "auto_generated")
    profile_name = _sanitize_profile_name(raw_name)
    generated_profile["profile_name"] = profile_name
    try:
        save_profile(profile_name, generated_profile)
    except Exception as exc:
        return json.dumps({"error": f"Failed to save profile: {exc}"})

    return json.dumps(
        {
            "status": "profile_generated",
            "profile_name": profile_name,
            "company": generated_profile.get("company", ""),
            "sheet_name": generated_profile.get("sheet_name", ""),
            "products": [p.get("name", "") for p in generated_profile.get("products", [])],
            "categories_mapped": len(generated_profile.get("category_mapping", {})),
            "message": (
                f"MappingProfile '{profile_name}' generated and saved. "
                "You can now use parse_ldi_sheet with this profile or 'auto'."
            ),
        }
    )


def _normalize_generated_profile(profile: dict) -> dict:
    """Normalize LLM-generated profile to expected format.

    LLMs sometimes produce shorthand forms:
    - column_mapping values as int instead of {"index": int}
    - column_mapping values as {"index": int} without "header"

    This function normalizes these to the canonical format.
    """
    if not isinstance(profile, dict):
        return profile

    # Normalize column_mapping (shared logic with mapping_profiles module)
    from lci_ignite.data.mapping_profiles import _normalize_column_mapping

    profile = _normalize_column_mapping(profile)

    # Ensure products have default fields
    products = profile.get("products", [])
    if isinstance(products, list):
        for p in products:
            if isinstance(p, dict):
                p.setdefault("total_energy_mj", 0)
                p.setdefault("fu_unit_factor", 0)
                p.setdefault("output_unit", "")

    return profile


def _validate_generated_profile(profile: dict) -> list[str]:
    """Validate a generated MappingProfile. Returns list of error messages."""
    errors: list[str] = []

    if not isinstance(profile, dict):
        return ["Profile must be a JSON object"]

    # Required top-level fields
    for field in ("profile_name", "sheet_name", "column_mapping"):
        if field not in profile:
            errors.append(f"Missing required field: {field}")

    # column_mapping must have core fields
    col_map = profile.get("column_mapping", {})
    if isinstance(col_map, dict):
        required_cols = {"process", "category", "flow_name", "direction", "unit"}
        missing_cols = required_cols - set(col_map.keys())
        if missing_cols:
            errors.append(f"column_mapping missing required fields: {missing_cols}")

        # Each column spec must have an index
        for key, spec in col_map.items():
            if isinstance(spec, dict) and "index" not in spec:
                errors.append(f"column_mapping['{key}'] missing 'index'")
    else:
        errors.append("column_mapping must be a dict")

    # Products must be a list of dicts with 'name'
    products = profile.get("products", [])
    if not isinstance(products, list):
        errors.append("products must be a list")
    else:
        for i, p in enumerate(products):
            if not isinstance(p, dict):
                errors.append(f"products[{i}] must be a dict")
            elif "name" not in p:
                errors.append(f"products[{i}] missing 'name'")

    return errors


# ── Data Processing Tools (7 NEW) ──

ANALYZE_EXCEL_STRUCTURE_SCHEMA = {
    "type": "function",
    "function": {
        "name": "analyze_excel_structure",
        "description": (
            "Extract Excel file structure for column mapping. Returns sheet names, "
            "headers, sample rows, dimensions, detected units and categories. "
            "No LLM call -- pure metadata extraction (Layer 2)."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Path to the Excel file (.xlsx)",
                },
            },
            "required": ["file_path"],
        },
    },
}


def analyze_excel_structure(file_path: str) -> str:
    """Extract Excel structure for LLM mapping (Layer 2).

    Returns JSON with sheet_names, headers, sample_rows, dimensions,
    detected_units, detected_categories. Also checks for saved MappingProfile match.
    """
    from lci_ignite.data.excel_profile import ExcelProfile
    from lci_ignite.data.mapping_profiles import match_profile

    try:
        profile = ExcelProfile.extract(file_path)
    except (FileNotFoundError, ValueError) as exc:
        return f"Error: {exc}"

    # Check if a saved profile matches
    matched = match_profile(profile)
    if matched:
        profile["matched_profile"] = matched.get("profile_name", "unknown")
        profile["message"] = (
            f"Found matching profile: {matched.get('profile_name')}. "
            "You can use parse_ldi_sheet with this profile."
        )
    else:
        profile["matched_profile"] = None
        profile["message"] = (
            "No matching profile found. Use generate_mapping_profile to auto-create one, "
            "or use parse_ldi_sheet with 'auto' which will trigger auto-generation."
        )

    return json.dumps(profile, indent=2, default=str)


PARSE_LDI_SHEET_SCHEMA = {
    "type": "function",
    "function": {
        "name": "parse_ldi_sheet",
        "description": (
            "Parse an LDI Master Excel sheet using a MappingProfile. "
            "Returns structured LCI data in Standard Schema format with flows, "
            "helper data, and summary. Use analyze_excel_structure first to identify the profile."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Path to the Excel file",
                },
                "profile_name": {
                    "type": "string",
                    "description": (
                        "Name of saved MappingProfile to load, "
                        "or 'auto' to auto-detect from file structure"
                    ),
                },
            },
            "required": ["file_path", "profile_name"],
        },
    },
}


def parse_ldi_sheet(file_path: str, profile_name: str) -> str:
    """Parse LDI Master sheet using a MappingProfile.

    Returns JSON Standard LCI Schema with flows, helper_data, products, summary.
    """
    from lci_ignite.data.excel_profile import ExcelProfile
    from lci_ignite.data.mapping_profiles import load_profile, match_profile
    from lci_ignite.data.sources.excel_lci import ExcelLCISource

    profile = None
    if profile_name == "auto":
        try:
            excel_profile = ExcelProfile.extract(file_path)
            profile = match_profile(excel_profile)
            if profile is None:
                # Layer 3: auto-generate profile via LLM
                logger.info("No matching profile found, triggering LLM auto-mapping...")
                gen_result = generate_mapping_profile(file_path)
                gen_data = json.loads(gen_result)
                if "error" in gen_data:
                    return (
                        f"Error: No matching profile found and auto-generation failed: "
                        f"{gen_data['error']}"
                    )
                # Load the newly generated profile
                generated_name = gen_data.get("profile_name", "")
                if not generated_name:
                    return "Error: Profile was generated but has no name."
                profile = load_profile(generated_name)
        except Exception as exc:
            return f"Error extracting profile: {exc}"
    else:
        try:
            profile = load_profile(profile_name)
        except FileNotFoundError:
            return f"Error: Profile '{profile_name}' not found."

    source = ExcelLCISource(path=file_path, profile=profile)
    try:
        raw_data = source.extract()
    except ValueError as exc:
        return f"Error parsing file: {exc}"

    # Store full data in pipeline state for downstream tools
    _store_pipeline(raw_data.content)

    # Return only summary (not full 366+ flows) to keep LLM context small
    summary = raw_data.content.get("summary", {})
    products = raw_data.content.get("products", [])
    return json.dumps(
        {
            "status": "parsed",
            "total_flows": summary.get("total_flows", 0),
            "categories": summary.get("categories", []),
            "processes": summary.get("processes", []),
            "products": products,
            "skipped_rows": summary.get("skipped_rows", 0),
            "message": (
                f"Parsed {summary.get('total_flows', 0)} flows across "
                f"{len(summary.get('categories', []))} categories. "
                "Data stored in pipeline. Use 'auto' for data parameter in next tools."
            ),
        },
        indent=2,
        default=str,
    )


AGGREGATE_FLOWS_SCHEMA = {
    "type": "function",
    "function": {
        "name": "aggregate_flows",
        "description": (
            "Merge duplicate flows that share the same category, flow_name, and unit. "
            "Sums numeric fields (amount, per_product_*, fu_*) and combines process names. "
            "Uses pipeline data automatically."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "data": {
                    "type": "string",
                    "description": "Use 'auto' to chain from previous pipeline step (default)",
                    "default": "auto",
                },
            },
            "required": [],
        },
    },
}


def aggregate_flows(data: str = "auto") -> str:
    """Merge duplicate flows that share the same (category, flow_name, unit).

    For each group:
    - Sums: amount, per_product_*, fu_*
    - Combines unique process names (sorted, comma-separated)
    - Keeps direction and original_category from first flow
    """
    parsed, err = _resolve_data(data)
    if err:
        return err
    flows = _extract_flows(parsed)

    from collections import OrderedDict

    groups: OrderedDict[tuple[str, str, str], dict] = OrderedDict()
    for flow in flows:
        key = (
            flow.get("category", ""),
            flow.get("flow_name", ""),
            flow.get("unit", ""),
        )
        if key not in groups:
            # First occurrence -- copy all fields as base
            groups[key] = dict(flow)
            # Track process names as a set for dedup
            groups[key]["_processes"] = {flow.get("process", "")} - {""}
        else:
            merged = groups[key]
            # Sum amount
            merged["amount"] = merged.get("amount", 0) + flow.get("amount", 0)
            # Sum per_product_* and fu_* fields
            for k, v in flow.items():
                if (k.startswith("per_product_") or k.startswith("fu_")) and isinstance(
                    v, (int, float)
                ):
                    merged[k] = merged.get(k, 0) + v
            # Collect process names
            proc = flow.get("process", "")
            if proc:
                merged["_processes"].add(proc)
            # Collect unique values for extra string fields
            _CORE_FIELDS = {
                "category",
                "original_category",
                "flow_name",
                "process",
                "direction",
                "amount",
                "unit",
                "_processes",
            }
            for k, v in flow.items():
                if k in _CORE_FIELDS or k.startswith("per_product_") or k.startswith("fu_"):
                    continue
                if not isinstance(v, str) or not v.strip():
                    continue
                existing = merged.get(k)
                if existing is None:
                    merged[k] = v
                elif isinstance(existing, str) and v.strip() not in existing:
                    merged[k] = f"{existing}, {v.strip()}"

    # Finalize: combine process names, remove temp field
    total_before = len(flows)
    aggregated: list[dict] = []
    for merged in groups.values():
        processes = merged.pop("_processes", set())
        if processes:
            merged["process"] = ", ".join(sorted(processes))
        aggregated.append(merged)

    total_after = len(aggregated)
    duplicates_merged = total_before - total_after

    # Build result carrying forward pipeline metadata
    result: dict[str, Any] = {"flows": aggregated}
    if isinstance(parsed, dict):
        for key in ("products", "helper_data", "summary", "fu_mode", "fu_unit_labels"):
            if key in parsed:
                result[key] = parsed[key]
    # Update summary counts if present
    if "summary" in result and isinstance(result["summary"], dict):
        result["summary"]["total_flows"] = total_after

    _store_pipeline(result)
    return json.dumps(
        {
            "status": "aggregated",
            "total_before": total_before,
            "total_after": total_after,
            "duplicates_merged": duplicates_merged,
            "message": (
                f"Merged {duplicates_merged} duplicate flows: "
                f"{total_before} \u2192 {total_after} unique flows."
            ),
        },
        indent=2,
    )


APPLY_UNIT_CONVERSIONS_SCHEMA = {
    "type": "function",
    "function": {
        "name": "apply_unit_conversions",
        "description": (
            "Apply unit conversions to parsed LCI data. Converts units like "
            "ton->kg, barrel->L, m3->L. Uses pipeline data automatically."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "data": {
                    "type": "string",
                    "description": "Use 'auto' to chain from previous pipeline step (default)",
                    "default": "auto",
                },
                "conversions": {
                    "type": "string",
                    "description": (
                        "JSON string of conversion rules array "
                        '[{"from_unit": "ton", "to_unit": "kg", "factor": 1000}]. '
                        "Default: empty (no conversions)"
                    ),
                    "default": "[]",
                },
            },
            "required": [],
        },
    },
}


def _extract_flows(parsed: Any) -> list[dict]:
    """Extract flows list from either a list or dict with 'flows' key.

    Tools accept data in two formats:
      - A flat list of flow dicts: [{...}, {...}]
      - A dict with a 'flows' key: {"flows": [{...}, ...], ...}
    This helper normalizes both to a plain list.
    """
    if isinstance(parsed, list):
        return parsed
    if isinstance(parsed, dict) and "flows" in parsed:
        return parsed["flows"]
    return parsed


def apply_unit_conversions(data: str = "auto", conversions: str = "[]") -> str:
    """Apply unit conversions to parsed LCI flows.

    Always applies baseline conversions from lci_schema.py (STANDARD_CATEGORIES)
    first, then applies any additional profile-specific rules on top.
    This ensures mixed units (barrel + liter + m3) within a category are
    always normalized to the category's default unit.
    """
    from lci_ignite.data.lci_schema import STANDARD_CATEGORIES

    parsed, err = _resolve_data(data)
    if err:
        return err
    flows = _extract_flows(parsed)
    try:
        profile_rules = json.loads(conversions)
    except json.JSONDecodeError as exc:
        return f"Error: Invalid conversions JSON - {exc}"

    # Build baseline conversion rules from lci_schema.py
    baseline_rules = _build_baseline_conversions(STANDARD_CATEGORIES)

    # Merge: profile rules override baseline for same from_unit+category
    merged_rules = _merge_conversion_rules(baseline_rules, profile_rules)

    converted = 0
    for flow in flows:
        cat = flow.get("category", "")
        unit = flow.get("unit", "").strip()
        amount = flow.get("amount", 0.0)

        if not unit or amount == 0:
            continue

        for rule in merged_rules:
            from_unit = rule["from_unit"]
            to_unit = rule["to_unit"]
            factor = rule["factor"]
            applies_to = rule.get("applies_to", [])

            if unit.lower() == from_unit.lower() and (not applies_to or cat in applies_to):
                flow["amount"] = amount * factor
                flow["unit"] = to_unit
                flow["original_amount"] = amount
                flow["original_unit"] = unit
                converted += 1
                break

    result = {
        "flows": flows,
        "conversions_applied": converted,
        "total_flows": len(flows),
    }
    # Carry forward products from pipeline state
    if isinstance(parsed, dict) and "products" in parsed:
        result["products"] = parsed["products"]
    _store_pipeline(result)
    return json.dumps(
        {
            "status": "converted",
            "conversions_applied": converted,
            "total_flows": len(flows),
            "baseline_rules": len(baseline_rules),
            "profile_rules": len(profile_rules),
            "message": f"Applied {converted} unit conversions across {len(flows)} flows.",
        },
        indent=2,
    )


def _build_baseline_conversions(
    categories: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    """Build conversion rules from STANDARD_CATEGORIES definitions.

    Reads the unit_conversions dict from each category and converts to
    the flat rule format: {from_unit, to_unit, factor, applies_to}.
    """
    rules: list[dict[str, Any]] = []
    for cat_name, cat_info in categories.items():
        default_unit = cat_info.get("default_unit")
        unit_convs = cat_info.get("unit_conversions")
        if not unit_convs or not default_unit:
            continue
        for from_unit, factor in unit_convs.items():
            if not isinstance(factor, (int, float)):
                continue  # skip special rules like "multiply_by_study_over_lifetime"
            rules.append(
                {
                    "from_unit": from_unit,
                    "to_unit": default_unit,
                    "factor": factor,
                    "applies_to": [cat_name],
                }
            )
    return rules


def _merge_conversion_rules(
    baseline: list[dict[str, Any]],
    profile: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Merge baseline and profile conversion rules.

    Profile rules override baseline for the same (from_unit, category) pair.
    """
    # Index profile rules by (from_unit_lower, category)
    profile_keys: set[tuple[str, str]] = set()
    for rule in profile:
        from_lower = rule.get("from_unit", "").lower()
        for cat in rule.get("applies_to", [""]):
            profile_keys.add((from_lower, cat))

    # Add baseline rules that are NOT overridden by profile
    merged = list(profile)
    for rule in baseline:
        from_lower = rule["from_unit"].lower()
        applies_to = rule.get("applies_to", [])
        # Check if any of this rule's categories are overridden
        overridden = any((from_lower, cat) in profile_keys for cat in applies_to)
        if not overridden:
            merged.append(rule)

    return merged


CALCULATE_FUNCTIONAL_UNIT_SCHEMA = {
    "type": "function",
    "function": {
        "name": "calculate_functional_unit",
        "description": (
            "Calculate per-MJ functional unit values for each product. "
            "Uses pipeline data and product definitions automatically."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "data": {
                    "type": "string",
                    "description": "Use 'auto' to chain from previous pipeline step (default)",
                    "default": "auto",
                },
                "products": {
                    "type": "string",
                    "description": "Use 'auto' to use products from parsed data (default)",
                    "default": "auto",
                },
                "fu_mode": {
                    "type": "string",
                    "enum": ["per_mj", "per_output_unit"],
                    "description": (
                        "FU mode. 'per_mj' (default PROPER) = divide by total energy MJ. "
                        "'per_output_unit' = divide by total output quantity "
                        "(per barrel, per MMSCF, per ton)."
                    ),
                    "default": "per_mj",
                },
            },
            "required": [],
        },
    },
}


def calculate_functional_unit(
    data: str = "auto",
    products: str = "auto",
    fu_mode: str = "per_mj",
) -> str:
    """Calculate functional unit values for each product.

    Modes:
        per_mj (default): divide per-product amount by total_energy_mj
        per_output_unit: divide per-product amount by total output quantity
            (total_energy_mj / fu_unit_factor = output in barrels, MMSCF, etc.)
    """
    parsed, err = _resolve_data(data)
    if err:
        return err
    flows = _extract_flows(parsed)

    # Products can come from pipeline state or explicit parameter
    if products == "auto":
        pipeline = _read_pipeline()
        if pipeline and "products" in pipeline:
            product_list = pipeline["products"]
        else:
            return "Error: No product definitions. Provide products parameter."
    else:
        try:
            product_list = json.loads(products)
        except json.JSONDecodeError as exc:
            return f"Error: Invalid products JSON - {exc}"

    if not product_list:
        return "Error: No product definitions provided"

    # Build per-product divisor and unit labels based on fu_mode
    fu_unit_labels: dict[str, str] = {}
    divisors: dict[str, float] = {}
    for product in product_list:
        name = product["name"]
        energy_mj = product.get("total_energy_mj", 0)
        fu_unit_factor = product.get("fu_unit_factor", 0)
        output_unit = product.get("output_unit", "")

        if fu_mode == "per_output_unit" and fu_unit_factor and energy_mj:
            # total output quantity = total_energy_mj / fu_unit_factor
            divisors[name] = energy_mj / fu_unit_factor
            fu_unit_labels[name] = output_unit or "unit"
        else:
            # Default: per MJ
            divisors[name] = energy_mj
            fu_unit_labels[name] = "MJ"

    for flow in flows:
        for product in product_list:
            name = product["name"]
            divisor = divisors.get(name, 0)
            if divisor == 0:
                continue

            per_product_key = f"per_product_{name}"
            per_product_amount = flow.get(per_product_key, 0.0)
            amount = flow.get("amount", 0.0)

            fu_key = f"fu_per_mj_{name}"
            if per_product_amount:
                flow[fu_key] = per_product_amount / divisor
            elif amount:
                # Fallback: when per-product column is empty/zero, use total amount
                flow[fu_key] = amount / divisor
            else:
                flow[fu_key] = 0.0

            pct_key = f"pct_{name}"
            flow[pct_key] = 0.0

    # Calculate percentage per category per product
    from collections import defaultdict

    cat_totals: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for flow in flows:
        cat = flow.get("category", "")
        for product in product_list:
            name = product["name"]
            fu_key = f"fu_per_mj_{name}"
            cat_totals[cat][name] += abs(flow.get(fu_key, 0.0))

    for flow in flows:
        cat = flow.get("category", "")
        for product in product_list:
            name = product["name"]
            fu_key = f"fu_per_mj_{name}"
            pct_key = f"pct_{name}"
            total = cat_totals[cat][name]
            if total > 0:
                flow[pct_key] = round(abs(flow.get(fu_key, 0.0)) / total * 100, 2)

    result = {
        "flows": flows,
        "products": product_list,
        "fu_mode": fu_mode,
        "fu_unit_labels": fu_unit_labels,
        "total_flows": len(flows),
    }
    _store_pipeline(result)

    mode_desc = "per MJ" if fu_mode == "per_mj" else "per output unit"
    return json.dumps(
        {
            "status": "fu_calculated",
            "products": [p["name"] for p in product_list],
            "fu_mode": fu_mode,
            "fu_unit_labels": fu_unit_labels,
            "total_flows": len(flows),
            "message": (
                f"Calculated FU values ({mode_desc}) for {len(product_list)} products"
                f" across {len(flows)} flows."
            ),
        },
        indent=2,
    )


SELECT_PARETO_ITEMS_SCHEMA = {
    "type": "function",
    "function": {
        "name": "select_pareto_items",
        "description": (
            "80/20 Pareto selection per LDI Category. "
            "Sorts by Total FU Amount (descending), keeps individual rows "
            "until 80% cumulative contribution is reached, aggregates the "
            "remaining ~20% into a single 'Others' row. "
            "Uses pipeline data automatically."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "data": {
                    "type": "string",
                    "description": "Use 'auto' to chain from previous pipeline step (default)",
                    "default": "auto",
                },
                "threshold": {
                    "type": "number",
                    "description": (
                        "Cumulative % threshold (0-100). Rows are kept until "
                        "this threshold is reached. Default: 80 (standard Pareto)"
                    ),
                    "default": 80,
                },
                "top_n": {
                    "type": "integer",
                    "description": (
                        "Optional hard cap on items per category. "
                        "If set, overrides threshold-based selection. Default: 0 (disabled)"
                    ),
                    "default": 0,
                },
                "others_label": {
                    "type": "string",
                    "description": "Label for aggregated remainder (default: 'Others')",
                    "default": "Others",
                },
            },
            "required": [],
        },
    },
}


def select_pareto_items(
    data: str = "auto",
    threshold: float = 80,
    top_n: int = 0,
    others_label: str = "Others",
    # Keep backward compat for old callers using lainnya_label
    lainnya_label: str | None = None,
) -> str:
    """80/20 Pareto selection per LDI Category.

    Algorithm:
    1. Sort items in each category by Total FU Amount (descending)
    2. Calculate cumulative % contribution
    3. Keep individual rows until threshold (default 80%) is reached
    4. Aggregate the rest into a single 'Others' row

    If top_n > 0, uses fixed top-N instead of threshold (backward compat).
    """
    parsed, err = _resolve_data(data)
    if err:
        return err
    flows = _extract_flows(parsed)

    label = lainnya_label if lainnya_label is not None else others_label

    from collections import defaultdict

    by_category: dict[str, list[dict]] = defaultdict(list)
    for flow in flows:
        cat = flow.get("category", "Unknown")
        by_category[cat].append(flow)

    # Detect sort key: prefer FU amount > per_product > amount
    sort_key = _detect_pareto_sort_key(flows)

    selected: list[dict] = []
    stats: dict[str, dict] = {}

    for cat, cat_flows in by_category.items():
        sorted_flows = sorted(
            cat_flows,
            key=lambda f: abs(_get_sort_value(f, sort_key)),
            reverse=True,
        )

        if top_n > 0:
            # Legacy mode: fixed top-N
            top_items = sorted_flows[:top_n]
            rest_items = sorted_flows[top_n:]
        else:
            # 80/20 mode: cumulative % threshold
            total = sum(abs(_get_sort_value(f, sort_key)) for f in sorted_flows)
            if total == 0:
                # All zero amounts — keep all
                top_items = sorted_flows
                rest_items = []
            else:
                cumulative = 0.0
                split_idx = len(sorted_flows)  # default: keep all
                for i, f in enumerate(sorted_flows):
                    cumulative += abs(_get_sort_value(f, sort_key))
                    pct = (cumulative / total) * 100
                    if pct >= threshold:
                        split_idx = i + 1  # include this row
                        break
                top_items = sorted_flows[:split_idx]
                rest_items = sorted_flows[split_idx:]

        # Add percentage to each kept item
        cat_total = sum(abs(_get_sort_value(f, sort_key)) for f in sorted_flows)
        for f in top_items:
            if cat_total > 0:
                f["pareto_pct"] = round(abs(_get_sort_value(f, sort_key)) / cat_total * 100, 2)
            else:
                f["pareto_pct"] = 0.0

        selected.extend(top_items)

        if rest_items:
            others_amount = sum(f.get("amount", 0) for f in rest_items)
            unit = rest_items[0].get("unit", "")
            others_entry: dict[str, Any] = {
                "category": cat,
                "flow_name": label,
                "amount": others_amount,
                "unit": unit,
                "direction": rest_items[0].get("direction", ""),
                "process": "",
                "is_aggregated": True,
                "aggregated_count": len(rest_items),
            }
            # Sum per_product and fu fields
            for key in rest_items[0]:
                if key.startswith("per_product_") or key.startswith("fu_"):
                    others_entry[key] = sum(f.get(key, 0) for f in rest_items)
            # Calculate Others percentage
            if cat_total > 0:
                others_sort_val = sum(abs(_get_sort_value(f, sort_key)) for f in rest_items)
                others_entry["pareto_pct"] = round(others_sort_val / cat_total * 100, 2)
            else:
                others_entry["pareto_pct"] = 0.0
            selected.append(others_entry)

        stats[cat] = {
            "total_items": len(cat_flows),
            "selected": len(top_items),
            "aggregated": len(rest_items),
            "sort_key": sort_key,
        }

    result = {
        "flows": selected,
        "stats": stats,
        "total_selected": len(selected),
    }
    # Carry forward products, fu_mode, fu_unit_labels from pipeline
    pipeline = _read_pipeline()
    if pipeline:
        for key in ("products", "fu_mode", "fu_unit_labels"):
            if key in pipeline:
                result[key] = pipeline[key]
    _store_pipeline(result)

    mode_desc = f"top_n={top_n}" if top_n > 0 else f"threshold={threshold}%"
    return json.dumps(
        {
            "status": "pareto_selected",
            "total_selected": len(selected),
            "categories": len(stats),
            "sort_key": sort_key,
            "mode": mode_desc,
            "stats": stats,
            "message": (
                f"Pareto {mode_desc}: {len(selected)} items across "
                f"{len(stats)} categories (sorted by {sort_key})."
            ),
        },
        indent=2,
    )


def _detect_pareto_sort_key(flows: list[dict]) -> str:
    """Detect the best sort key for Pareto ranking.

    Priority: fu_per_mj_* > per_product_* > amount
    """
    if not flows:
        return "amount"
    sample = flows[0]
    # Check for FU keys (calculated by calculate_functional_unit)
    fu_keys = [k for k in sample if k.startswith("fu_per_mj_")]
    if fu_keys:
        return fu_keys[0]  # Use first product's FU
    # Check for per-product keys
    pp_keys = [k for k in sample if k.startswith("per_product_")]
    if pp_keys:
        return pp_keys[0]
    return "amount"


def _get_sort_value(flow: dict, sort_key: str) -> float:
    """Get the numeric value for Pareto sorting."""
    val = flow.get(sort_key, 0)
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


VALIDATE_DATA_QUALITY_SCHEMA = {
    "type": "function",
    "function": {
        "name": "validate_data_quality",
        "description": (
            "Check for known data quality issues in LCI data. Uses pipeline data automatically."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "data": {
                    "type": "string",
                    "description": "Use 'auto' to chain from previous pipeline step (default)",
                    "default": "auto",
                },
            },
            "required": [],
        },
    },
}


def validate_data_quality(data: str = "auto") -> str:
    """Check for known data quality issues in LCI data."""
    parsed, err = _resolve_data(data)
    if err:
        return err
    flows = _extract_flows(parsed)

    issues: list[dict[str, str]] = []

    # Check 1: Duplicate values across different emission categories
    emission_flows = [f for f in flows if "Emisi" in f.get("category", "")]
    emission_amounts: dict[float, list[str]] = {}
    for flow in emission_flows:
        amt = round(flow.get("amount", 0), 5)
        if amt != 0:
            cat = flow.get("category", "")
            emission_amounts.setdefault(amt, []).append(cat)

    for amt, cats in emission_amounts.items():
        unique_cats = set(cats)
        if len(unique_cats) > 1:
            issues.append(
                {
                    "severity": "CRITICAL",
                    "category": ", ".join(sorted(unique_cats)),
                    "description": (
                        f"Identical value {amt} found across different emission categories: "
                        f"{', '.join(sorted(unique_cats))}. Possible copy-paste error."
                    ),
                    "suggestion": "Verify source data for these emission categories.",
                }
            )

    # Check 2: Suspiciously small values that may be missing *1000 conversion
    for flow in flows:
        cat = flow.get("category", "")
        unit = flow.get("unit", "")
        amount = flow.get("amount", 0)
        if unit == "kg" and 0 < abs(amount) < 1 and "Emisi" in cat:
            issues.append(
                {
                    "severity": "CRITICAL",
                    "category": cat,
                    "description": (
                        f"Flow '{flow.get('flow_name', '')}' has very small kg value ({amount}). "
                        "May be missing *1000 conversion from ton."
                    ),
                    "suggestion": "Check if original unit was ton and conversion was missed.",
                }
            )

    # Check 3: Zero amounts
    zero_count = sum(1 for f in flows if f.get("amount", 0) == 0)
    if zero_count > 0:
        issues.append(
            {
                "severity": "MINOR",
                "category": "All",
                "description": f"{zero_count} flows have zero amounts.",
                "suggestion": "Verify if these are intentional or missing data.",
            }
        )

    # Check 4: Negative amounts
    negative_flows = [f for f in flows if f.get("amount", 0) < 0]
    if negative_flows:
        names = ", ".join(f.get("flow_name", "") for f in negative_flows[:3])
        issues.append(
            {
                "severity": "MODERATE",
                "category": "Various",
                "description": f"{len(negative_flows)} flows have negative amounts: {names}",
                "suggestion": "Negative values are unusual in LCI. Check direction assignment.",
            }
        )

    # Push callout if critical issues found
    if any(i["severity"] == "CRITICAL" for i in issues):
        callout_content = "**Data Quality Issues Detected:**\n"
        for issue in issues:
            if issue["severity"] == "CRITICAL":
                callout_content += f"- [{issue['severity']}] {issue['description']}\n"
        item: dict[str, Any] = {
            "calloutContent": callout_content,
            "variant": "warning",
            "title": "Data Quality Alert",
        }
        items = _get_render_list()
        items.append(item)

    result = {
        "issues": issues,
        "total_issues": len(issues),
        "critical": len([i for i in issues if i["severity"] == "CRITICAL"]),
        "moderate": len([i for i in issues if i["severity"] == "MODERATE"]),
        "minor": len([i for i in issues if i["severity"] == "MINOR"]),
    }
    return json.dumps(result, indent=2)


# ── Emission sub-section classification ──

_EMISSION_POLLUTANTS = [
    "CO2",
    "CH4",
    "CO",
    "NOx",
    "N2O",
    "SOx",
    "Particulate Material",
    "nmVOC",
    "TOC",
]

# Regex patterns for emission classification.
# Order matters: longer/more-specific patterns first to avoid "CO2" matching "CO".
# "PM" is accepted as alias for "Particulate Material".
_EMISSION_PATTERNS: list[tuple[str, str]] = [
    (r"(?i)\bCO2\b", "CO2"),
    (r"(?i)\bCH4\b", "CH4"),
    (r"(?i)\bNOx\b", "NOx"),
    (r"(?i)\bN2O\b", "N2O"),
    (r"(?i)\bSOx\b", "SOx"),
    (r"(?i)\b(?:PM|Particulate)\b", "Particulate Material"),
    (r"(?i)\bnmVOC\b", "nmVOC"),
    (r"(?i)\bTOC\b", "TOC"),
    # CO must come after CO2 to avoid false matches
    (r"(?i)\bCO\b(?!2)", "CO"),
]


def _classify_emission(flow_name: str) -> str | None:
    """Classify an emission flow into a pollutant sub-section.

    Returns pollutant name (e.g. "CO2", "Particulate Material") or None.
    """
    name = flow_name.strip()
    for pattern, pollutant in _EMISSION_PATTERNS:
        if re.search(pattern, name):
            return pollutant
    return None


def _build_emission_subsections(
    emisi_flows: list[dict],
    products: list[dict],
    top_n: int = 5,
) -> dict[str, list[dict]]:
    """Build emission sub-sections from Emisi Udara flows.

    Returns dict mapping section name -> list of flow rows, including:
    - "Emisi Udara" summary (8 aggregate rows, one per pollutant)
    - "Emisi CO2", "Emisi CH4", ... detail sections (top N + Lainnya + Total)
    """
    from collections import defaultdict

    # Classify each flow
    by_pollutant: dict[str, list[dict]] = defaultdict(list)
    for flow in emisi_flows:
        pollutant = _classify_emission(flow.get("flow_name", ""))
        if pollutant:
            by_pollutant[pollutant].append(flow)
        else:
            # Unclassified emissions go into a catch-all
            by_pollutant["_other"].append(flow)

    sections: dict[str, list[dict]] = {}

    # Summary section: one aggregate row per pollutant
    summary_flows: list[dict] = []
    for pollutant in _EMISSION_POLLUTANTS:
        p_flows = by_pollutant.get(pollutant, [])
        if not p_flows:
            continue
        total_amount = sum(f.get("amount", 0.0) for f in p_flows)
        unit = p_flows[0].get("unit", "kg")
        agg: dict[str, Any] = {
            "flow_name": f"Total {pollutant}",
            "amount": total_amount,
            "unit": unit,
            "direction": "output",
            "category": "Emisi Udara",
            "process": "",
        }
        # Aggregate FU values
        for product in products:
            name = product.get("name", "")
            fu_key = f"fu_per_mj_{name}"
            agg[fu_key] = sum(f.get(fu_key, 0.0) for f in p_flows)
        summary_flows.append(agg)

    if summary_flows:
        sections["Emisi Udara"] = summary_flows

    # Detail sections per pollutant
    for pollutant in _EMISSION_POLLUTANTS:
        p_flows = by_pollutant.get(pollutant, [])
        if not p_flows:
            continue

        section_name = f"Emisi {pollutant}"
        sorted_flows = sorted(p_flows, key=lambda f: abs(f.get("amount", 0)), reverse=True)

        detail_flows: list[dict] = []
        top_items = sorted_flows[:top_n]
        rest_items = sorted_flows[top_n:]

        detail_flows.extend(top_items)

        if rest_items:
            lainnya_amount = sum(f.get("amount", 0) for f in rest_items)
            unit = rest_items[0].get("unit", "kg")
            lainnya: dict[str, Any] = {
                "flow_name": "Lainnya",
                "amount": lainnya_amount,
                "unit": unit,
                "direction": "output",
                "category": section_name,
                "process": "",
                "is_aggregated": True,
                "aggregated_count": len(rest_items),
            }
            for product in products:
                name = product.get("name", "")
                fu_key = f"fu_per_mj_{name}"
                lainnya[fu_key] = sum(f.get(fu_key, 0.0) for f in rest_items)
            detail_flows.append(lainnya)

        sections[section_name] = detail_flows

    return sections


BUILD_PROPER_IO_TABLE_SCHEMA = {
    "type": "function",
    "function": {
        "name": "build_proper_io_table",
        "description": (
            "Build a full PROPER-format IO Table with 11 columns. "
            "Uses pipeline data and product config automatically."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "data": {
                    "type": "string",
                    "description": "Use 'auto' to chain from previous pipeline step (default)",
                    "default": "auto",
                },
                "config": {
                    "type": "string",
                    "description": "Use 'auto' to derive from pipeline data (default)",
                    "default": "auto",
                },
            },
            "required": [],
        },
    },
}


def build_proper_io_table(data: str = "auto", config: str = "auto") -> str:
    """Build full PROPER-format IO Table.

    Columns per product: Item | Total | Unit | Product FU/{unit} | Unit | %
    Flows are already aggregated by material (no process/area column).
    """
    from lci_ignite.data.lci_schema import IO_TABLE_SECTION_ORDER

    parsed, err = _resolve_data(data)
    if err:
        return err
    flows = _extract_flows(parsed)

    # Config can come from pipeline state or explicit parameter
    pipeline = _read_pipeline()
    if config == "auto":
        products_from_pipeline = []
        if pipeline and "products" in pipeline:
            product_names = pipeline["products"]
            if product_names and isinstance(product_names[0], str):
                products_from_pipeline = [{"name": n} for n in product_names]
            elif product_names and isinstance(product_names[0], dict):
                products_from_pipeline = product_names
        table_config = {"products": products_from_pipeline, "title": "IO Table PROPER"}
    else:
        try:
            table_config = json.loads(config)
        except json.JSONDecodeError as exc:
            return f"Error: Invalid config JSON - {exc}"

    products = table_config.get("products", [])
    title = table_config.get("title", "IO Table")

    # Read FU unit labels from pipeline
    fu_unit_labels: dict[str, str] = {}
    if pipeline:
        fu_unit_labels = pipeline.get("fu_unit_labels", {})

    # Build headers (no Process/Area column — flows are already aggregated by material)
    headers = ["Input/Output", "Total", "Unit"]
    for product in products:
        name = product.get("name", "Product")
        fu_label = fu_unit_labels.get(name, "MJ")
        headers.extend([f"{name} FU/{fu_label}", "Unit", "%"])

    # Group flows by category
    from collections import defaultdict

    by_category: dict[str, list[dict]] = defaultdict(list)
    for flow in flows:
        cat = flow.get("category", "Unknown")
        by_category[cat].append(flow)

    # Build emission sub-sections from Emisi Udara flows
    emisi_flows = by_category.pop("Emisi Udara", [])
    emission_sections: dict[str, list[dict]] = {}
    if emisi_flows:
        emission_sections = _build_emission_subsections(emisi_flows, products)
        # Merge emission sub-sections into by_category for rendering
        for section_name, section_flows in emission_sections.items():
            by_category[section_name] = section_flows

    rows: list[list[str]] = []
    section_count = 0

    def _render_section(section: str, section_flows: list[dict]) -> None:
        nonlocal section_count
        if not section_flows:
            return

        section_count += 1

        # Section header row
        hdr_row = [f"**{section}**"] + [""] * (len(headers) - 1)
        rows.append(hdr_row)

        sorted_flows = sorted(section_flows, key=lambda f: abs(f.get("amount", 0)), reverse=True)

        section_total = 0.0
        for flow in sorted_flows:
            amount = flow.get("amount", 0.0)
            section_total += amount
            unit = flow.get("unit", "")
            flow_name = flow.get("flow_name", "")

            row = [flow_name, _fmt_number(amount), unit]

            for product in products:
                name = product.get("name", "")
                fu_key = f"fu_per_mj_{name}"
                pct_key = f"pct_{name}"
                fu_val = flow.get(fu_key, 0.0)
                pct_val = flow.get(pct_key, 0.0)
                fu_label = fu_unit_labels.get(name, "MJ")
                fu_unit = f"{unit}/{fu_label}" if unit else ""

                row.extend(
                    [
                        _fmt_number(fu_val),
                        fu_unit,
                        f"{pct_val:.1f}" if pct_val else "",
                    ]
                )

            rows.append(row)

        # Section total row
        if len(sorted_flows) > 1:
            total_row = [
                f"Total {section}",
                _fmt_number(section_total),
                sorted_flows[0].get("unit", ""),
            ]
            for product in products:
                name = product.get("name", "")
                fu_key = f"fu_per_mj_{name}"
                fu_total = sum(f.get(fu_key, 0.0) for f in sorted_flows)
                total_row.extend([_fmt_number(fu_total), "", "100.0"])
            rows.append(total_row)

    for section in IO_TABLE_SECTION_ORDER:
        _render_section(section, by_category.get(section, []))

    # Push to render items
    item = {"headers": headers, "rows": rows, "title": title}
    items = _get_render_list()
    items[:] = [
        i for i in items if not (i.get("title") == title and "headers" in i and "rows" in i)
    ]
    items.append(item)

    return (
        f"PROPER IO Table created: '{title}' with {section_count} sections, "
        f"{len(rows)} rows, {len(headers)} columns."
    )


def _fmt_number(value: float) -> str:
    """Format a number for IO Table display."""
    if value == 0:
        return "0"
    if abs(value) >= 1000:
        return f"{value:,.2f}"
    if abs(value) >= 1:
        return f"{value:.2f}"
    if abs(value) >= 0.01:
        return f"{value:.4f}"
    return f"{value:.6e}"


# ── Conversational Tools (4 NEW) ──

EXPLAIN_ANALYSIS_SCHEMA = {
    "type": "function",
    "function": {
        "name": "explain_analysis",
        "description": (
            "Explain analysis results by extracting relevant data context. "
            "Use when the user asks questions like 'kenapa CO2 paling tinggi?' "
            "or 'jelaskan emisi udara'. Returns structured data for the LLM to narrate."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "question": {
                    "type": "string",
                    "description": "The user's question about the analysis results",
                },
                "data": {
                    "type": "string",
                    "description": "Use 'auto' to read from pipeline (default)",
                    "default": "auto",
                },
            },
            "required": ["question"],
        },
    },
}


def explain_analysis(question: str, data: str = "auto") -> str:
    """Extract relevant pipeline data to answer a user question about results.

    Detects intent from question keywords and returns structured context:
    - Category drill-down (e.g. "kenapa CO2 paling tinggi?")
    - Top contributors per category
    - Process-level analysis
    """
    parsed, err = _resolve_data(data)
    if err:
        return err
    flows = _extract_flows(parsed)

    if not flows:
        return json.dumps({"error": "No pipeline data available. Run the analysis pipeline first."})

    q_lower = question.lower()

    # Detect category focus from question
    target_category = None
    category_keywords = {
        "co2": "Emisi CO2",
        "ch4": "Emisi CH4",
        "nox": "Emisi NOx",
        "n2o": "Emisi N2O",
        "sox": "Emisi SOx",
        "pm": "Emisi Particulate Material",
        "nmvoc": "Emisi nmVOC",
        "toc": "Emisi TOC",
        "co": "Emisi CO",
        "emisi udara": "Emisi Udara",
        "bahan baku": "Bahan Baku",
        "energi": "Energi",
        "air": "Air",
        "limbah": "Limbah B3",
    }
    for keyword, cat in category_keywords.items():
        if keyword in q_lower:
            target_category = cat
            break

    # Filter flows
    if target_category:
        relevant = [f for f in flows if f.get("category", "") == target_category]
    else:
        relevant = flows

    if not relevant:
        # Fallback: try partial match
        for f in flows:
            cat = f.get("category", "").lower()
            if any(kw in cat for kw in q_lower.split()):
                relevant.append(f)

    if not relevant:
        relevant = flows

    # Sort by amount descending
    relevant = sorted(relevant, key=lambda f: abs(f.get("amount", 0)), reverse=True)

    # Compute totals and percentages — return ALL fields per flow
    total_amount = sum(abs(f.get("amount", 0)) for f in relevant)
    result_flows = []
    for f in relevant[:30]:
        amount = f.get("amount", 0)
        pct = (abs(amount) / total_amount * 100) if total_amount else 0
        entry = dict(f)  # Copy all fields
        entry["percentage"] = round(pct, 2)
        result_flows.append(entry)

    top = result_flows[0] if result_flows else None
    result = {
        "question": question,
        "target_category": target_category,
        "total_amount": round(total_amount, 4),
        "unit": top["unit"] if top else "",
        "flow_count": len(relevant),
        "top_flows": result_flows,
        "top_contributor": top["flow_name"] if top else "",
        "top_contributor_pct": top["percentage"] if top else 0,
        "top_process": top["process"] if top else "",
    }

    # Push a callout with the key finding
    if top:
        callout = (
            f"**{target_category or 'Analysis'}**: "
            f"{top['flow_name']} is the largest contributor "
            f"({top['percentage']:.1f}% of total {round(total_amount, 2)} {top['unit']})"
        )
        item: dict[str, Any] = {
            "calloutContent": callout,
            "variant": "info",
            "title": "Analysis Insight",
        }
        items = _get_render_list()
        items.append(item)

    return json.dumps(result, indent=2)


COMPARE_PRODUCTS_SCHEMA = {
    "type": "function",
    "function": {
        "name": "compare_products",
        "description": (
            "Compare products side-by-side across categories. "
            "Use when the user asks 'bandingkan Gas vs Minyak' or similar. "
            "Returns per-category FU/MJ comparison with deltas."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "metric": {
                    "type": "string",
                    "description": (
                        "Comparison focus: 'all' for all categories, "
                        "or a specific category name (e.g. 'Emisi CO2')"
                    ),
                    "default": "all",
                },
                "data": {
                    "type": "string",
                    "description": "Use 'auto' to read from pipeline (default)",
                    "default": "auto",
                },
            },
            "required": [],
        },
    },
}


def compare_products(metric: str = "all", data: str = "auto") -> str:
    """Compare products side-by-side with per-category FU/MJ values."""
    parsed, err = _resolve_data(data)
    if err:
        return err
    flows = _extract_flows(parsed)

    if not flows:
        return json.dumps({"error": "No pipeline data available."})

    # Detect product names from FU keys
    product_names = []
    sample = flows[0] if flows else {}
    for key in sample:
        if key.startswith("fu_per_mj_"):
            product_names.append(key.replace("fu_per_mj_", ""))

    if not product_names:
        return json.dumps({"error": "No FU/MJ data found. Run calculate_functional_unit first."})

    # Filter by metric
    if metric != "all":
        target_flows = [f for f in flows if f.get("category", "") == metric]
        if not target_flows:
            # Partial match
            target_flows = [f for f in flows if metric.lower() in f.get("category", "").lower()]
        if not target_flows:
            target_flows = flows
    else:
        target_flows = flows

    # Build per-category comparison
    from collections import defaultdict

    cat_data: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for f in target_flows:
        cat = f.get("category", "Unknown")
        for pname in product_names:
            fu_key = f"fu_per_mj_{pname}"
            cat_data[cat][pname] += abs(f.get(fu_key, 0.0))

    comparison_rows = []
    for cat in sorted(cat_data.keys()):
        row: dict[str, Any] = {"category": cat}
        values = []
        for pname in product_names:
            val = round(cat_data[cat][pname], 6)
            row[pname] = val
            values.append(val)
        # Delta and ratio (only for 2-product case)
        if len(product_names) == 2 and values[1] != 0:
            row["delta"] = round(values[0] - values[1], 6)
            row["ratio"] = round(values[0] / values[1], 4)
        comparison_rows.append(row)

    # Read FU unit labels from pipeline for headers
    pipeline = _read_pipeline()
    fu_unit_labels: dict[str, str] = {}
    if pipeline:
        fu_unit_labels = pipeline.get("fu_unit_labels", {})

    # Push ObTable with dynamic FU unit labels
    display_names = []
    for pname in product_names:
        fu_label = fu_unit_labels.get(pname, "MJ")
        display_names.append(f"{pname} (/{fu_label})")
    headers = ["Category"] + display_names
    if len(product_names) == 2:
        headers += ["Delta", "Ratio"]
    table_rows = []
    for r in comparison_rows:
        table_row = [r["category"]]
        for pname in product_names:
            table_row.append(_fmt_number(r[pname]))
        if len(product_names) == 2:
            table_row.append(_fmt_number(r.get("delta", 0)))
            table_row.append(str(r.get("ratio", "")))
        table_rows.append(table_row)

    table_item = {
        "headers": headers,
        "rows": table_rows,
        "title": f"Product Comparison: {metric}",
    }
    items = _get_render_list()
    items.append(table_item)

    result = {
        "products": product_names,
        "metric": metric,
        "categories_compared": len(comparison_rows),
        "comparison": comparison_rows,
    }
    return json.dumps(result, indent=2)


REVISE_PIPELINE_SCHEMA = {
    "type": "function",
    "function": {
        "name": "revise_pipeline",
        "description": (
            "Re-run part of the pipeline with revised parameters. "
            "Use when the user asks 'ubah top N jadi 10' or 'recalculate'. "
            "Actions: 'set_top_n' (rerun Pareto selection), 'recalculate_fu' (rerun FU calc)."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "action": {
                    "type": "string",
                    "enum": ["set_top_n", "recalculate_fu"],
                    "description": "Which pipeline step to re-run",
                },
                "value": {
                    "type": "integer",
                    "description": "New parameter value (e.g. top_n=10)",
                },
                "data": {
                    "type": "string",
                    "description": "Use 'auto' to read from pipeline (default)",
                    "default": "auto",
                },
                "fu_mode": {
                    "type": "string",
                    "enum": ["per_mj", "per_output_unit"],
                    "description": (
                        "FU mode for recalculate_fu action. "
                        "'per_mj' = per MJ (default), "
                        "'per_output_unit' = per output unit (barrel, MMSCF, etc.)"
                    ),
                    "default": "per_mj",
                },
            },
            "required": ["action", "value"],
        },
    },
}


def revise_pipeline(action: str, value: int, data: str = "auto", fu_mode: str = "per_mj") -> str:
    """Re-run a pipeline step with revised parameters."""
    parsed, err = _resolve_data(data)
    if err:
        return err
    flows = _extract_flows(parsed)
    if not flows:
        return json.dumps({"error": "No pipeline data available."})

    if action == "set_top_n":
        before_count = len(flows)
        result_str = select_pareto_items(data="auto", top_n=value)
        after_pipeline = _read_pipeline()
        after_count = len(_extract_flows(after_pipeline)) if after_pipeline else 0
        return json.dumps(
            {
                "action": "set_top_n",
                "new_value": value,
                "before_flow_count": before_count,
                "after_flow_count": after_count,
                "pareto_result": json.loads(result_str),
                "message": f"Re-ran Pareto selection with top_n={value}. "
                f"Flows: {before_count} -> {after_count}.",
            },
            indent=2,
        )

    elif action == "recalculate_fu":
        result_str = calculate_functional_unit(data="auto", products="auto", fu_mode=fu_mode)
        return json.dumps(
            {
                "action": "recalculate_fu",
                "fu_result": json.loads(result_str),
                "message": "Recalculated functional unit values.",
            },
            indent=2,
        )

    return json.dumps({"error": f"Unknown action: {action}"})


EXPORT_FILTERED_SCHEMA = {
    "type": "function",
    "function": {
        "name": "export_filtered",
        "description": (
            "Export a filtered subset of the IO Table to Excel. "
            "Use when the user asks 'export hanya section Emisi Udara' or similar. "
            "Accepts section name filters."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "sections": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": (
                        "Section names to include (e.g. ['Emisi CO2', 'Emisi CH4']). "
                        "Shortcuts: 'inputs' = all input sections, "
                        "'outputs' = all output sections, "
                        "'emissions' = all Emisi sections."
                    ),
                },
                "filename": {
                    "type": "string",
                    "description": "Output filename (default: io_table_filtered.xlsx)",
                    "default": "io_table_filtered.xlsx",
                },
                "data": {
                    "type": "string",
                    "description": "Use 'auto' to read from pipeline (default)",
                    "default": "auto",
                },
            },
            "required": ["sections"],
        },
    },
}


def export_filtered(
    sections: list[str],
    filename: str = "io_table_filtered.xlsx",
    data: str = "auto",
) -> str:
    """Export a filtered subset of pipeline flows to Excel."""
    from lci_ignite.data.lci_schema import STANDARD_CATEGORIES

    parsed, err = _resolve_data(data)
    if err:
        return err
    flows = _extract_flows(parsed)
    if not flows:
        return json.dumps({"error": "No pipeline data available."})

    # Expand shortcuts
    expanded: set[str] = set()
    for s in sections:
        s_lower = s.lower()
        if s_lower == "inputs":
            for cat, info in STANDARD_CATEGORIES.items():
                if info.get("direction") == "input":
                    expanded.add(cat)
        elif s_lower == "outputs":
            for cat, info in STANDARD_CATEGORIES.items():
                if info.get("direction") == "output":
                    expanded.add(cat)
        elif s_lower == "emissions":
            for cat in STANDARD_CATEGORIES:
                if "Emisi" in cat:
                    expanded.add(cat)
            # Also include generated emission sub-sections
            for f in flows:
                cat = f.get("category", "")
                if "Emisi" in cat:
                    expanded.add(cat)
        else:
            expanded.add(s)

    # Filter flows
    filtered = [f for f in flows if f.get("category", "") in expanded]
    if not filtered:
        return json.dumps(
            {
                "error": f"No flows found matching sections: {list(expanded)}",
                "available_categories": sorted(set(f.get("category", "") for f in flows)),
            }
        )

    # Temporarily store filtered data and export
    original_pipeline = _read_pipeline()
    filtered_data = {"flows": filtered}

    # Copy products from pipeline
    if original_pipeline and "products" in original_pipeline:
        filtered_data["products"] = original_pipeline["products"]

    _store_pipeline(filtered_data)
    result = export_to_xlsx(data="auto", filename=filename, title="IO Table (Filtered)")

    # Restore original pipeline
    if original_pipeline:
        _store_pipeline(original_pipeline)

    return json.dumps(
        {
            "exported_sections": sorted(expanded),
            "filtered_flow_count": len(filtered),
            "total_flow_count": len(flows),
            "filename": filename,
            "export_result": result,
        },
        indent=2,
    )


# ── Query / Filter Tool ──

QUERY_FLOWS_SCHEMA = {
    "type": "function",
    "function": {
        "name": "query_flows",
        "description": (
            "Query and filter pipeline flows by any field. "
            "Use when the user asks to filter, search, or aggregate by any column, "
            "including extra columns from the original Excel file like: "
            "data_source, pic, notes, review_status, produced_from, "
            "material_composition, sample_size, abbreviation, parameter, etc. "
            "Examples: 'tampilkan flow yang data_source = Measured', "
            "'siapa PIC untuk emisi udara?', "
            "'filter review_status = C', "
            "'group by data_source'."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "field": {
                    "type": "string",
                    "description": (
                        "Field name to filter/query by (snake_case). "
                        "Common fields: category, flow_name, process, direction, unit, "
                        "data_source, pic, review_status, produced_from, "
                        "material_composition, notes, sample_size, abbreviation, parameter, "
                        "is_amount_balanced, data_source_reference"
                    ),
                },
                "value": {
                    "type": "string",
                    "description": (
                        "Value to match (case-insensitive partial match). "
                        "Use '*' or omit for all values (group-by mode)."
                    ),
                    "default": "*",
                },
                "mode": {
                    "type": "string",
                    "enum": ["filter", "group", "unique", "list_fields"],
                    "description": (
                        "'filter' = show flows matching field=value, "
                        "'group' = aggregate amounts by field values, "
                        "'unique' = list all unique values for a field, "
                        "'list_fields' = show all available fields in pipeline data"
                    ),
                    "default": "filter",
                },
                "data": {
                    "type": "string",
                    "description": "Use 'auto' to read from pipeline (default)",
                    "default": "auto",
                },
            },
            "required": ["field"],
        },
    },
}


def query_flows(field: str, value: str = "*", mode: str = "filter", data: str = "auto") -> str:
    """Query and filter pipeline flows by any field.

    Supports filtering, grouping, and listing unique values for any field
    including extra columns extracted from the original Excel file.
    """
    parsed, err = _resolve_data(data)
    if err:
        return err
    flows = _extract_flows(parsed)
    if not flows:
        return json.dumps({"error": "No pipeline data available. Run the analysis pipeline first."})

    # Mode: list_fields — show all available field names
    if mode == "list_fields":
        all_keys: set[str] = set()
        for f in flows:
            all_keys.update(f.keys())
        return json.dumps(
            {
                "available_fields": sorted(all_keys),
                "total_flows": len(flows),
                "message": "Use any of these fields with query_flows to filter or aggregate.",
            },
            indent=2,
        )

    field_lower = field.lower().strip()

    # Mode: unique — list all distinct values for a field
    if mode == "unique":
        values: list[str] = []
        for f in flows:
            val = _find_field(f, field_lower)
            if val is not None:
                val_str = str(val).strip()
                if val_str and val_str not in values:
                    values.append(val_str)
        return json.dumps(
            {
                "field": field,
                "unique_values": sorted(values),
                "count": len(values),
            },
            indent=2,
        )

    # Mode: group — aggregate amounts by field values
    if mode == "group":
        groups: dict[str, dict] = {}
        for f in flows:
            val = _find_field(f, field_lower)
            group_key = str(val).strip() if val is not None else "(empty)"
            if group_key not in groups:
                groups[group_key] = {"count": 0, "total_amount": 0.0, "categories": set()}
            groups[group_key]["count"] += 1
            groups[group_key]["total_amount"] += abs(f.get("amount", 0))
            groups[group_key]["categories"].add(f.get("category", ""))

        result_groups = []
        for key, info in sorted(groups.items(), key=lambda x: x[1]["total_amount"], reverse=True):
            result_groups.append(
                {
                    "value": key,
                    "flow_count": info["count"],
                    "total_amount": round(info["total_amount"], 4),
                    "categories": sorted(info["categories"]),
                }
            )

        # Push table render item
        headers = [field, "Flow Count", "Total Amount", "Categories"]
        rows = [
            [
                g["value"],
                str(g["flow_count"]),
                _fmt_number(g["total_amount"]),
                ", ".join(g["categories"][:5]),
            ]
            for g in result_groups[:20]
        ]
        items = _get_render_list()
        items.append(
            {
                "title": f"Group by: {field}",
                "headers": headers,
                "rows": rows,
            }
        )

        return json.dumps(
            {
                "field": field,
                "groups": result_groups,
                "total_groups": len(result_groups),
            },
            indent=2,
        )

    # Mode: filter — show flows matching field=value
    if value == "*":
        matched = flows
    else:
        value_lower = value.lower().strip()
        matched = []
        for f in flows:
            val = _find_field(f, field_lower)
            if val is not None and value_lower in str(val).lower():
                matched.append(f)

    if not matched:
        # Suggest available values
        available = set()
        for f in flows:
            val = _find_field(f, field_lower)
            if val is not None:
                available.add(str(val).strip())
        return json.dumps(
            {
                "field": field,
                "filter_value": value,
                "matched": 0,
                "available_values": sorted(available)[:20],
                "message": f"No flows found with {field} matching '{value}'.",
            },
            indent=2,
        )

    # Return ALL fields per flow so LLM can reason over complete data
    result_flows = matched[:50]

    # Push table render item (UI shows key columns only)
    headers = ["Flow", "Category", "Amount", "Unit", "Process", field]
    rows = [
        [
            f.get("flow_name", ""),
            f.get("category", ""),
            _fmt_number(f.get("amount", 0)),
            f.get("unit", ""),
            f.get("process", ""),
            str(_find_field(f, field_lower) or ""),
        ]
        for f in result_flows
    ]
    items = _get_render_list()
    items.append(
        {
            "title": (
                f"Flows where {field} = '{value}'" if value != "*" else f"All flows ({field})"
            ),
            "headers": headers,
            "rows": rows,
        }
    )

    return json.dumps(
        {
            "field": field,
            "filter_value": value,
            "matched": len(matched),
            "total_flows": len(flows),
            "flows": result_flows,
        },
        indent=2,
        default=str,
    )


def _find_field(flow: dict, field_lower: str) -> Any:
    """Find a field value in a flow dict (case-insensitive key match)."""
    # Exact match first
    if field_lower in flow:
        return flow[field_lower]
    # Case-insensitive search
    for k, v in flow.items():
        if k.lower() == field_lower:
            return v
    return None


# ── Chart Generation Tools (auto-read pipeline data) ──

GENERATE_CATEGORY_CHART_SCHEMA = {
    "type": "function",
    "function": {
        "name": "generate_category_chart",
        "description": (
            "Generate a horizontal bar chart showing total amounts per IO Table category. "
            "Auto-reads pipeline data — no JSON parameter needed. "
            "Groups flows by category and direction (input vs output), "
            "sorted by PROPER section order."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "title": {
                    "type": "string",
                    "description": "Chart title (default: 'IO Table by Category')",
                    "default": "IO Table by Category",
                },
                "data": {
                    "type": "string",
                    "description": "Use 'auto' to read from pipeline (default)",
                    "default": "auto",
                },
            },
            "required": [],
        },
    },
}


def generate_category_chart(
    title: str = "IO Table by Category",
    data: str = "auto",
) -> str:
    """Generate a bar chart of total amounts per IO Table category.

    Auto-reads pipeline flows, groups by category + direction, sums amounts,
    and pushes a Recharts-compatible bar chart render item.
    """
    from lci_ignite.data.lci_schema import IO_TABLE_SECTION_ORDER, STANDARD_CATEGORIES

    parsed, err = _resolve_data(data)
    if err:
        return err
    flows = _extract_flows(parsed)
    if not flows:
        return json.dumps({"error": "No pipeline data available. Run parse_ldi_sheet first."})

    # Sum amounts per category, split by direction
    from collections import defaultdict

    cat_input: dict[str, float] = defaultdict(float)
    cat_output: dict[str, float] = defaultdict(float)

    for flow in flows:
        cat = flow.get("category", "")
        amount = abs(float(flow.get("amount", 0)))
        direction = flow.get("direction", "")
        if not direction:
            cat_info = STANDARD_CATEGORIES.get(cat)
            direction = cat_info["direction"] if cat_info else "input"
        if direction == "input":
            cat_input[cat] += amount
        else:
            cat_output[cat] += amount

    # Build chart data sorted by PROPER section order
    all_cats = set(cat_input.keys()) | set(cat_output.keys())
    ordered = [c for c in IO_TABLE_SECTION_ORDER if c in all_cats]
    # Append any categories not in standard order
    ordered += sorted(c for c in all_cats if c not in ordered)

    chart_data = []
    for cat in ordered:
        entry: dict[str, Any] = {"category": cat}
        inp = cat_input.get(cat, 0)
        out = cat_output.get(cat, 0)
        if inp > 0:
            entry["input"] = round(inp, 4)
        if out > 0:
            entry["output"] = round(out, 4)
        chart_data.append(entry)

    if not chart_data:
        return json.dumps({"error": "No category data to chart."})

    item: dict[str, Any] = {
        "type": "bar",
        "title": title,
        "data": chart_data,
        "options": {
            "xKey": "category",
            "series": [
                {"dataKey": "input", "name": "Input"},
                {"dataKey": "output", "name": "Output"},
            ],
        },
    }
    items = _get_render_list()
    items[:] = [i for i in items if not (i.get("title") == title and "data" in i)]
    items.append(item)

    return json.dumps(
        {
            "status": "chart_created",
            "title": title,
            "categories": len(chart_data),
            "message": f"Category chart '{title}' created with {len(chart_data)} categories.",
        }
    )


GENERATE_EMISSION_CHART_SCHEMA = {
    "type": "function",
    "function": {
        "name": "generate_emission_chart",
        "description": (
            "Generate a pie chart showing emission breakdown by pollutant type. "
            "Auto-reads pipeline data, filters 'Emisi Udara' flows, classifies "
            "by pollutant (CO2, CH4, NOx, etc.), and shows total kg + percentage."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "title": {
                    "type": "string",
                    "description": "Chart title (default: 'Emission Breakdown')",
                    "default": "Emission Breakdown",
                },
                "data": {
                    "type": "string",
                    "description": "Use 'auto' to read from pipeline (default)",
                    "default": "auto",
                },
            },
            "required": [],
        },
    },
}


def generate_emission_chart(
    title: str = "Emission Breakdown",
    data: str = "auto",
) -> str:
    """Generate a pie chart of emission breakdown by pollutant.

    Auto-reads pipeline flows, filters Emisi Udara category, classifies
    each flow into pollutant types via _classify_emission(), sums per
    pollutant, and pushes a pie chart render item.
    """
    parsed, err = _resolve_data(data)
    if err:
        return err
    flows = _extract_flows(parsed)
    if not flows:
        return json.dumps({"error": "No pipeline data available. Run parse_ldi_sheet first."})

    # Filter Emisi Udara flows
    emisi_flows = [f for f in flows if f.get("category", "") == "Emisi Udara"]
    if not emisi_flows:
        return json.dumps(
            {"error": "No 'Emisi Udara' flows found. Check pipeline data categories."}
        )

    # Classify and sum per pollutant
    from collections import defaultdict

    pollutant_totals: dict[str, float] = defaultdict(float)
    for flow in emisi_flows:
        pollutant = _classify_emission(flow.get("flow_name", ""))
        label = pollutant or "Other"
        pollutant_totals[label] += abs(float(flow.get("amount", 0)))

    grand_total = sum(pollutant_totals.values())
    if grand_total == 0:
        return json.dumps({"error": "Total emission amount is zero."})

    # Build pie chart data sorted by value descending
    chart_data = []
    for name, value in sorted(pollutant_totals.items(), key=lambda x: x[1], reverse=True):
        pct = round(value / grand_total * 100, 2)
        chart_data.append({"name": name, "value": round(value, 4), "percentage": pct})

    item: dict[str, Any] = {
        "type": "pie",
        "title": title,
        "data": chart_data,
    }
    items = _get_render_list()
    items[:] = [i for i in items if not (i.get("title") == title and "data" in i)]
    items.append(item)

    return json.dumps(
        {
            "status": "chart_created",
            "title": title,
            "pollutants": len(chart_data),
            "total_kg": round(grand_total, 4),
            "message": (
                f"Emission chart '{title}' created with {len(chart_data)} pollutant types "
                f"(total {round(grand_total, 2)} kg)."
            ),
        }
    )


GENERATE_PRODUCT_CHART_SCHEMA = {
    "type": "function",
    "function": {
        "name": "generate_product_chart",
        "description": (
            "Generate a grouped bar chart comparing products across IO Table categories. "
            "Auto-reads pipeline data and products list. "
            "Shows per-product amounts grouped by category."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "title": {
                    "type": "string",
                    "description": "Chart title (default: 'Product Comparison by Category')",
                    "default": "Product Comparison by Category",
                },
                "data": {
                    "type": "string",
                    "description": "Use 'auto' to read from pipeline (default)",
                    "default": "auto",
                },
            },
            "required": [],
        },
    },
}


def generate_product_chart(
    title: str = "Product Comparison by Category",
    data: str = "auto",
) -> str:
    """Generate a grouped bar chart comparing products across categories.

    Auto-reads pipeline flows and products list, groups by category,
    sums per_product_{name} amounts, and pushes a grouped bar chart.
    """
    from lci_ignite.data.lci_schema import IO_TABLE_SECTION_ORDER

    parsed, err = _resolve_data(data)
    if err:
        return err
    flows = _extract_flows(parsed)
    if not flows:
        return json.dumps({"error": "No pipeline data available. Run parse_ldi_sheet first."})

    # Detect product names from per_product_ keys
    product_names: list[str] = []
    sample = flows[0] if flows else {}
    for key in sorted(sample.keys()):
        if key.startswith("per_product_"):
            product_names.append(key.replace("per_product_", ""))

    if not product_names:
        # Fallback: check pipeline-level products
        pipeline = _read_pipeline()
        if pipeline and "products" in pipeline:
            prod_list = pipeline["products"]
            if prod_list:
                if isinstance(prod_list[0], str):
                    product_names = prod_list
                elif isinstance(prod_list[0], dict):
                    product_names = [p["name"] for p in prod_list]

    if not product_names:
        return json.dumps({"error": "No product data found. Run calculate_functional_unit first."})

    # Sum per_product amounts per category
    from collections import defaultdict

    cat_products: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for flow in flows:
        cat = flow.get("category", "")
        for pname in product_names:
            key = f"per_product_{pname}"
            cat_products[cat][pname] += abs(float(flow.get(key, 0)))

    # Order by PROPER section order
    all_cats = set(cat_products.keys())
    ordered = [c for c in IO_TABLE_SECTION_ORDER if c in all_cats]
    ordered += sorted(c for c in all_cats if c not in ordered)

    chart_data = []
    for cat in ordered:
        entry: dict[str, Any] = {"category": cat}
        for pname in product_names:
            entry[pname] = round(cat_products[cat][pname], 4)
        chart_data.append(entry)

    if not chart_data:
        return json.dumps({"error": "No category data to chart."})

    series = [{"dataKey": pname, "name": pname} for pname in product_names]

    item: dict[str, Any] = {
        "type": "bar",
        "title": title,
        "data": chart_data,
        "options": {
            "xKey": "category",
            "series": series,
        },
    }
    items = _get_render_list()
    items[:] = [i for i in items if not (i.get("title") == title and "data" in i)]
    items.append(item)

    return json.dumps(
        {
            "status": "chart_created",
            "title": title,
            "categories": len(chart_data),
            "products": product_names,
            "message": (
                f"Product chart '{title}' created with {len(chart_data)} categories "
                f"and {len(product_names)} products."
            ),
        }
    )
