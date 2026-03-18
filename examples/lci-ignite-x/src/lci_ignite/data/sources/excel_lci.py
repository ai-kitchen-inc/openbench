"""Excel LCI data source -- generic parser for any company's LDI Master.

Uses a MappingProfile to parse arbitrary Excel LDI formats into the Standard
LCI Schema. When no profile is provided, extracts structural metadata for
LLM-based mapping (Layer 3).
"""

from __future__ import annotations

import hashlib
import logging
from pathlib import Path
from typing import Any

import openpyxl
from openbench.core.abstractions import DataSource, RawData

from lci_ignite.data.excel_profile import ExcelProfile
from lci_ignite.data.lci_schema import (
    EXCLUDED_LDI_CATEGORIES,
    HELPER_LDI_CATEGORIES,
    normalize_category,
    normalize_unit,
)

logger = logging.getLogger(__name__)

# Units that indicate a Co-Product helper row (not a real waste flow)
_CO_PRODUCT_HELPER_UNITS = frozenset(
    {
        "mj",
        "year",
        "unit well",
        "jumlah",
        "jumlah pekerjaan",
        "jumlah workover",
    }
)

# Flow names in Co-Product that map to Limbah Cair
_CO_PRODUCT_LIQUID_WASTE = frozenset(
    {
        "water produced",
        "water injection",
        "produced water",
        "injection water",
    }
)


class ExcelLCISource(DataSource):
    """Parse any Excel LDI file using a MappingProfile.

    Two modes:
        1. With profile: deterministic parsing -> Standard LCI Schema flows
        2. Without profile: structural extraction -> metadata for LLM mapping

    Args:
        path: Path to the Excel file (.xlsx).
        profile: MappingProfile dict (from mapping_profiles/ or LLM-generated).
        sheet_name: Override sheet name (default: from profile or first sheet).
    """

    def __init__(
        self,
        path: str | Path,
        profile: dict[str, Any] | None = None,
        sheet_name: str | None = None,
    ):
        self._path = Path(path)
        self._profile = profile
        self._sheet_name = sheet_name

    @property
    def source_type(self) -> str:
        return "excel_lci"

    @property
    def source_id(self) -> str:
        path_hash = hashlib.md5(str(self._path.resolve()).encode()).hexdigest()[:8]
        profile_name = self._profile.get("profile_name", "unknown") if self._profile else "none"
        return f"excel_lci_{profile_name}_{path_hash}"

    def get_metadata(self) -> dict[str, Any]:
        meta: dict[str, Any] = {
            "path": str(self._path),
            "format": "excel_lci",
        }
        if self._profile:
            meta["profile_name"] = self._profile.get("profile_name", "")
            meta["company"] = self._profile.get("company", "")
        if self._path.exists():
            meta["size_bytes"] = self._path.stat().st_size
        return meta

    def validate(self) -> bool:
        """Validate that the file exists and is a valid Excel file."""
        if not self._path.exists():
            logger.warning("File not found: %s", self._path)
            return False

        if self._path.suffix.lower() not in (".xlsx", ".xls"):
            logger.warning("Not an Excel file: %s", self._path.suffix)
            return False

        try:
            wb = openpyxl.load_workbook(str(self._path), read_only=True, data_only=True)
            sheet_name = self._resolve_sheet_name(wb)
            if sheet_name not in wb.sheetnames:
                logger.warning("Sheet '%s' not found in %s", sheet_name, self._path)
                wb.close()
                return False
            wb.close()
            return True
        except Exception as e:
            logger.warning("Failed to validate %s: %s", self._path, e)
            return False

    def extract(self) -> RawData:
        """Extract LCI data from Excel file.

        Returns:
            - With profile: RawData with Standard LCI Schema flows
            - Without profile: RawData with structural metadata for LLM mapping
        """
        if self._profile:
            return self._parse_with_profile()
        else:
            return self._extract_structure()

    def _resolve_sheet_name(self, wb: openpyxl.Workbook) -> str:
        """Determine which sheet to read."""
        if self._sheet_name:
            return self._sheet_name
        if self._profile and self._profile.get("sheet_name"):
            return self._profile["sheet_name"]
        return wb.sheetnames[0]

    # ------------------------------------------------------------------
    # Mode 1: Parse with MappingProfile (deterministic)
    # ------------------------------------------------------------------

    def _parse_with_profile(self) -> RawData:
        """Parse Excel LDI using a MappingProfile into Standard LCI Schema.

        The profile maps column indices to semantic fields (process, category,
        flow_name, direction, unit, scope_value, per_product amounts).
        """
        if not self.validate():
            raise ValueError(f"Invalid Excel file: {self._path}")

        profile = self._profile
        col_map = profile["column_mapping"]
        cat_map = profile.get("category_mapping", {})

        wb = openpyxl.load_workbook(str(self._path), data_only=True, read_only=True)
        sheet_name = self._resolve_sheet_name(wb)
        ws = wb[sheet_name]

        header_row = profile.get("header_row", 1)

        flows: list[dict[str, Any]] = []
        helper_data: dict[str, list[dict[str, Any]]] = {}
        categories_seen: set[str] = set()
        processes_seen: set[str] = set()
        skipped_rows = 0

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx <= header_row:
                continue

            row_list = list(row)
            if not row_list or all(v is None for v in row_list):
                continue

            # Extract fields using column mapping
            raw_category = self._get_cell(row_list, col_map.get("category"))
            if not raw_category:
                skipped_rows += 1
                continue

            raw_category = str(raw_category).strip()

            # Check if excluded or helper
            if raw_category in EXCLUDED_LDI_CATEGORIES:
                skipped_rows += 1
                continue

            if raw_category in HELPER_LDI_CATEGORIES:
                helper_entry = self._extract_helper_row(row_list, col_map, raw_category)
                if helper_entry:
                    helper_data.setdefault(raw_category, []).append(helper_entry)
                continue

            # Map LDI category to standard category
            standard_cat = cat_map.get(raw_category)
            if standard_cat is None:
                standard_cat = normalize_category(raw_category)
            if standard_cat is None:
                logger.debug("Unmapped category: %s", raw_category)
                skipped_rows += 1
                continue

            process = str(self._get_cell(row_list, col_map.get("process")) or "").strip()
            flow_name = str(self._get_cell(row_list, col_map.get("flow_name")) or "").strip()
            direction_raw = str(self._get_cell(row_list, col_map.get("direction")) or "").strip()
            direction = direction_raw.lower()
            unit = str(self._get_cell(row_list, col_map.get("unit")) or "").strip()
            scope_value = self._get_numeric(row_list, col_map.get("scope_value"))

            if direction not in ("input", "output"):
                skipped_rows += 1
                continue

            # ── Co-Product filtering ──
            if standard_cat == "Co-Product":
                resolved = self._resolve_co_product(flow_name, unit)
                if resolved is None:
                    skipped_rows += 1
                    continue
                standard_cat = resolved

            # ── Land splitting ──
            if standard_cat == "Lahan Digunakan":
                name_lower = flow_name.lower()
                if "ditransformasi" in name_lower or "transformed" in name_lower:
                    standard_cat = "Lahan Ditransformasi"

            flow: dict[str, Any] = {
                "category": standard_cat,
                "original_category": raw_category,
                "flow_name": flow_name,
                "process": process,
                "direction": direction,
                "amount": scope_value if scope_value is not None else 0.0,
                "unit": normalize_unit(unit),
            }

            # Per-product amounts (if present in profile)
            for product in profile.get("products", []):
                product_col = product.get("column")
                if product_col and product_col in col_map:
                    product_amount = self._get_numeric(row_list, col_map[product_col])
                    flow[f"per_product_{product['name']}"] = product_amount or 0.0

                fu_col = product.get("fu_column")
                if fu_col and fu_col in col_map:
                    fu_value = self._get_numeric(row_list, col_map[fu_col])
                    flow[f"fu_{product['name']}"] = fu_value or 0.0

            flows.append(flow)
            categories_seen.add(standard_cat)
            if process:
                processes_seen.add(process)

        wb.close()

        # Extract product definitions
        products = []
        for p in profile.get("products", []):
            products.append(
                {
                    "name": p["name"],
                    "total_energy_mj": p.get("total_energy_mj", 0),
                    "fu_unit_factor": p.get("fu_unit_factor", 0),
                    "output_unit": p.get("output_unit", ""),
                }
            )

        content: dict[str, Any] = {
            "flows": flows,
            "helper_data": helper_data,
            "products": products,
            "summary": {
                "total_flows": len(flows),
                "skipped_rows": skipped_rows,
                "categories": sorted(categories_seen),
                "processes": sorted(processes_seen),
            },
        }

        metadata = {
            "rows_parsed": len(flows) + skipped_rows,
            "flow_count": len(flows),
            "category_count": len(categories_seen),
            "process_count": len(processes_seen),
            "profile_name": profile.get("profile_name", ""),
            "sheet_name": sheet_name,
            "path": str(self._path),
        }

        return RawData(
            content=content,
            content_type="structured",
            metadata=metadata,
            source=self,
        )

    # ------------------------------------------------------------------
    # Mode 2: Extract structure for LLM mapping
    # ------------------------------------------------------------------

    def _extract_structure(self) -> RawData:
        """Extract Excel structural metadata for LLM-based mapping.

        Used when no MappingProfile exists. Returns enough information for
        the LLM to generate a MappingProfile.
        """
        profile = ExcelProfile.extract(self._path)

        content = {
            "mode": "structure_extraction",
            "excel_profile": profile,
            "message": (
                "No MappingProfile found for this Excel file. "
                "Use the excel_profile data to create a column mapping."
            ),
        }

        metadata = {
            "path": str(self._path),
            "sheet_count": len(profile["sheet_names"]),
            "sheet_names": profile["sheet_names"],
        }

        return RawData(
            content=content,
            content_type="excel_profile",
            metadata=metadata,
            source=self,
        )

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _get_cell(row: list[Any], col_spec: dict[str, Any] | None) -> Any:
        """Get cell value from row using column spec (with 0-based index)."""
        if col_spec is None:
            return None
        idx = col_spec.get("index")
        if idx is None or idx < 0 or idx >= len(row):
            return None
        return row[idx]

    @staticmethod
    def _get_numeric(row: list[Any], col_spec: dict[str, Any] | None) -> float | None:
        """Get numeric cell value, returning None if not numeric."""
        val = ExcelLCISource._get_cell(row, col_spec)
        if val is None:
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _extract_helper_row(
        row: list[Any],
        col_map: dict[str, Any],
        category: str,
    ) -> dict[str, Any] | None:
        """Extract a helper data row (lifetime values)."""
        flow_name = ExcelLCISource._get_cell(row, col_map.get("flow_name"))
        scope_value = ExcelLCISource._get_numeric(row, col_map.get("scope_value"))
        unit = ExcelLCISource._get_cell(row, col_map.get("unit"))

        if flow_name is None and scope_value is None:
            return None

        return {
            "category": category,
            "flow_name": str(flow_name).strip() if flow_name else "",
            "value": scope_value or 0.0,
            "unit": str(unit).strip() if unit else "",
        }

    @staticmethod
    def _resolve_co_product(flow_name: str, unit: str) -> str | None:
        """Resolve a Co-Product row to its actual standard category.

        Returns:
            - "Limbah Cair" for water-related co-products
            - None to skip helper/fuel rows
        """
        unit_lower = unit.lower().strip()
        name_lower = flow_name.lower().strip()

        # Skip helper rows (MJ, year, well counts, etc.)
        if unit_lower in _CO_PRODUCT_HELPER_UNITS:
            return None

        # Water Produced / Water Injection -> Limbah Cair
        if any(kw in name_lower for kw in _CO_PRODUCT_LIQUID_WASTE):
            return "Limbah Cair"

        # Fuel Gas co-products -> skip (already counted under Fuel Gas category)
        if "fuel gas" in name_lower or "gas" in name_lower:
            return None

        # Other Co-Product rows -> skip
        return None
