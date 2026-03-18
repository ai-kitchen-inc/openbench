"""ExcelProfile — deterministic Excel metadata extraction (Layer 2).

Extracts structural metadata from an Excel file without any LLM involvement.
The output is used by the LLM (Layer 3) to create a MappingProfile, or by
the profile matcher to find a saved MappingProfile.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)


class ExcelProfile:
    """Deterministic Excel metadata extractor.

    Reads an .xlsx file and produces a structural profile with sheet names,
    headers, sample rows, dimensions, detected units, and categories.
    """

    @staticmethod
    def extract(path: str | Path) -> dict[str, Any]:
        """Extract Excel structure profile for LLM mapping or profile matching.

        Args:
            path: Path to the Excel file (.xlsx).

        Returns:
            Dict with keys:
                - file_name: str
                - sheet_names: list[str]
                - sheets: dict[str, SheetProfile]

            Each SheetProfile has:
                - name: str
                - dimensions: {rows: int, cols: int}
                - headers: list[str | None]
                - header_row: int (1-based)
                - sample_rows: list[list[Any]] (first 5 data rows)
                - detected_units: list[str] (unique units found)
                - detected_categories: list[str] (unique categories found)
                - empty_columns: list[int] (0-based indices of empty cols)

        Raises:
            FileNotFoundError: If the file does not exist.
            ValueError: If the file is not a valid Excel file.
        """
        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {path}")

        if path.suffix.lower() not in (".xlsx", ".xls"):
            raise ValueError(f"Not an Excel file: {path.suffix}")

        try:
            wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        except Exception as exc:
            raise ValueError(f"Failed to open Excel file: {exc}") from exc

        profile: dict[str, Any] = {
            "file_name": path.name,
            "sheet_names": wb.sheetnames,
            "sheets": {},
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_profile = ExcelProfile._extract_sheet(ws)
            profile["sheets"][sheet_name] = sheet_profile

        wb.close()
        return profile

    @staticmethod
    def extract_sheet(path: str | Path, sheet_name: str | None = None) -> dict[str, Any]:
        """Extract profile for a single sheet.

        Args:
            path: Path to the Excel file.
            sheet_name: Name of the sheet. Uses the first sheet if None.

        Returns:
            SheetProfile dict.
        """
        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {path}")

        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        if sheet_name is None:
            sheet_name = wb.sheetnames[0]
        elif sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

        ws = wb[sheet_name]
        result = ExcelProfile._extract_sheet(ws)
        wb.close()
        return result

    @staticmethod
    def _extract_sheet(ws) -> dict[str, Any]:
        """Extract metadata from a single worksheet."""
        rows_data: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            rows_data.append(list(row))
            if len(rows_data) > 50:
                break

        if not rows_data:
            return {
                "name": ws.title,
                "dimensions": {"rows": 0, "cols": 0},
                "headers": [],
                "header_row": 0,
                "sample_rows": [],
                "detected_units": [],
                "detected_categories": [],
                "empty_columns": [],
            }

        # Find the header row (first row where majority of cells are non-empty strings)
        header_row_idx = ExcelProfile._find_header_row(rows_data)
        headers = [str(v).strip() if v is not None else None for v in rows_data[header_row_idx]]

        # Sample rows (up to 5 data rows after header)
        data_start = header_row_idx + 1
        sample_rows = []
        for row in rows_data[data_start : data_start + 5]:
            sample_rows.append([ExcelProfile._serialize_cell(v) for v in row])

        # Detect units and categories from data rows
        detected_units = ExcelProfile._detect_column_values(
            rows_data[data_start:], headers, ["unit", "satuan"]
        )
        detected_categories = ExcelProfile._detect_column_values(
            rows_data[data_start:], headers, ["category", "kategori", "ldi"]
        )

        # Find empty columns
        num_cols = len(headers)
        empty_columns = []
        for col_idx in range(num_cols):
            all_empty = all(
                row[col_idx] is None for row in rows_data[data_start:] if col_idx < len(row)
            )
            if all_empty:
                empty_columns.append(col_idx)

        # Count total rows (estimate from what we read + ws dimensions)
        total_rows = ws.max_row if ws.max_row else len(rows_data)

        return {
            "name": ws.title,
            "dimensions": {"rows": total_rows, "cols": num_cols},
            "headers": headers,
            "header_row": header_row_idx + 1,  # 1-based
            "sample_rows": sample_rows,
            "detected_units": sorted(set(detected_units)),
            "detected_categories": sorted(set(detected_categories)),
            "empty_columns": empty_columns,
        }

    @staticmethod
    def _find_header_row(rows: list[list[Any]], max_scan: int = 10) -> int:
        """Find the header row index (0-based).

        Heuristic: the first row where >50% of cells are non-empty strings.
        """
        for i, row in enumerate(rows[:max_scan]):
            if not row:
                continue
            non_empty = sum(1 for v in row if v is not None and str(v).strip())
            ratio = non_empty / len(row)
            if ratio > 0.5 and non_empty >= 3:
                # Check if values look like headers (mostly strings, not numbers)
                string_count = sum(
                    1 for v in row if v is not None and isinstance(v, str) and v.strip()
                )
                if string_count >= non_empty * 0.5:
                    return i
        return 0

    @staticmethod
    def _detect_column_values(
        data_rows: list[list[Any]],
        headers: list[str | None],
        keywords: list[str],
    ) -> list[str]:
        """Detect unique values from columns whose header matches keywords."""
        values: list[str] = []
        for col_idx, header in enumerate(headers):
            if header is None:
                continue
            header_lower = header.lower()
            if any(kw in header_lower for kw in keywords):
                for row in data_rows:
                    if col_idx < len(row) and row[col_idx] is not None:
                        val = str(row[col_idx]).strip()
                        if val and val not in values:
                            values.append(val)
        return values

    @staticmethod
    def _serialize_cell(value: Any) -> Any:
        """Serialize a cell value to JSON-safe format."""
        if value is None:
            return None
        if isinstance(value, (int, float, bool)):
            return value
        return str(value)
