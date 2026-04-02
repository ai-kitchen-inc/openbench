"""Tests for ExcelLCISource -- Excel LDI parser with MappingProfile."""

from pathlib import Path

import openpyxl

from lci_ignite.data.sources.excel_lci import ExcelLCISource

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

SIMPLE_PROFILE = {
    "profile_name": "test_profile",
    "company": "Test Corp",
    "sheet_name": "LDI Master",
    "header_row": 1,
    "column_mapping": {
        "process": {"index": 0, "header": "Process"},
        "category": {"index": 1, "header": "Category"},
        "flow_name": {"index": 2, "header": "Material"},
        "direction": {"index": 3, "header": "Direction"},
        "unit": {"index": 4, "header": "Unit"},
        "scope_value": {"index": 5, "header": "Amount"},
    },
    "category_mapping": {
        "Water": "Air",
        "Electricity": "Listrik",
        "Product": "Produk",
        "Air Emissions": "Emisi Udara",
    },
    "products": [],
    "unit_conversions": [],
}


def _create_ldi_xlsx(path: Path, sheet_name: str = "LDI Master", rows=None):
    """Create a test Excel file mimicking an LDI Master sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = ["Process", "Category", "Material", "Direction", "Unit", "Amount"]
    ws.append(headers)

    if rows is None:
        rows = [
            ["Well Operation", "Water", "Water Produced", "Input", "L", 1000.0],
            ["Well Operation", "Water", "Water Injection", "Input", "barrel", 500.0],
            ["NSOP", "Electricity", "Pompa", "Input", "kWh", 609586.18],
            ["Well Operation", "Product", "Minyak Bumi", "Output", "Barrel", 369113.5],
            ["NSOP", "Air Emissions", "CO2 Flaring", "Output", "ton", 19607.03],
        ]

    for row in rows:
        ws.append(row)

    wb.save(str(path))


def _create_ldi_xlsx_with_helpers(path: Path):
    """Create xlsx with helper categories (lifetime values)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LDI Master"

    headers = ["Process", "Category", "Material", "Direction", "Unit", "Amount"]
    ws.append(headers)

    ws.append(["Well Op", "Water", "Produced Water", "Input", "L", 5000.0])
    ws.append(["Well Op", "Projected Lifetime of Infrastructure", "Tank", "Input", "year", 20.0])
    ws.append(["Well Op", "Raw Material from Processes", "Recycled", "Input", "kg", 100.0])
    ws.append(["NSOP", "Electricity", "Pompa", "Input", "kWh", 1000.0])

    wb.save(str(path))


# ---------------------------------------------------------------------------
# Tests: Properties
# ---------------------------------------------------------------------------


class TestExcelLCISourceProperties:
    def test_source_type(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        _create_ldi_xlsx(xlsx)
        source = ExcelLCISource(xlsx, profile=SIMPLE_PROFILE)
        assert source.source_type == "excel_lci"

    def test_source_id_with_profile(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        _create_ldi_xlsx(xlsx)
        source = ExcelLCISource(xlsx, profile=SIMPLE_PROFILE)
        assert "test_profile" in source.source_id

    def test_source_id_without_profile(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        _create_ldi_xlsx(xlsx)
        source = ExcelLCISource(xlsx)
        assert "none" in source.source_id

    def test_metadata_with_profile(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        _create_ldi_xlsx(xlsx)
        source = ExcelLCISource(xlsx, profile=SIMPLE_PROFILE)
        meta = source.get_metadata()
        assert meta["format"] == "excel_lci"
        assert meta["profile_name"] == "test_profile"
        assert "size_bytes" in meta

    def test_metadata_without_profile(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        _create_ldi_xlsx(xlsx)
        source = ExcelLCISource(xlsx)
        meta = source.get_metadata()
        assert "profile_name" not in meta


# ---------------------------------------------------------------------------
# Tests: Validation
# ---------------------------------------------------------------------------


class TestExcelLCISourceValidation:
    def test_file_not_found(self):
        source = ExcelLCISource("/nonexistent/file.xlsx", profile=SIMPLE_PROFILE)
        assert source.validate() is False

    def test_not_excel_file(self, tmp_path):
        csv_file = tmp_path / "data.csv"
        csv_file.write_text("a,b,c")
        source = ExcelLCISource(csv_file, profile=SIMPLE_PROFILE)
        assert source.validate() is False

    def test_valid_file(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        _create_ldi_xlsx(xlsx)
        source = ExcelLCISource(xlsx, profile=SIMPLE_PROFILE)
        assert source.validate() is True

    def test_sheet_name_mismatch_falls_back(self, tmp_path):
        """When profile sheet doesn't exist, falls back to first sheet."""
        xlsx = tmp_path / "test.xlsx"
        _create_ldi_xlsx(xlsx, sheet_name="Other Sheet")
        source = ExcelLCISource(xlsx, profile=SIMPLE_PROFILE)
        # Profile expects "LDI Master" but file has "Other Sheet"
        # Now falls back to first sheet instead of failing
        assert source.validate() is True

    def test_sheet_name_override(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        _create_ldi_xlsx(xlsx, sheet_name="Custom")
        source = ExcelLCISource(xlsx, profile=SIMPLE_PROFILE, sheet_name="Custom")
        assert source.validate() is True

    def test_sheet_fallback_parses_correctly(self, tmp_path):
        """Fallback to first sheet still parses data correctly."""
        xlsx = tmp_path / "test.xlsx"
        _create_ldi_xlsx(xlsx, sheet_name="Actual Data")
        source = ExcelLCISource(xlsx, profile=SIMPLE_PROFILE)
        result = source.extract()
        assert result.content_type == "structured"
        flows = result.content["flows"]
        assert len(flows) == 5


# ---------------------------------------------------------------------------
# Tests: Parse with Profile
# ---------------------------------------------------------------------------


class TestParseWithProfile:
    def test_basic_parsing(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        _create_ldi_xlsx(xlsx)
        source = ExcelLCISource(xlsx, profile=SIMPLE_PROFILE)
        result = source.extract()

        assert result.content_type == "structured"
        flows = result.content["flows"]
        assert len(flows) == 5

    def test_category_mapping(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        _create_ldi_xlsx(xlsx)
        source = ExcelLCISource(xlsx, profile=SIMPLE_PROFILE)
        result = source.extract()

        categories = {f["category"] for f in result.content["flows"]}
        # "Water" -> "Air", "Electricity" -> "Listrik", etc.
        assert "Air" in categories
        assert "Listrik" in categories
        assert "Produk" in categories

    def test_direction_normalized(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        _create_ldi_xlsx(xlsx)
        source = ExcelLCISource(xlsx, profile=SIMPLE_PROFILE)
        result = source.extract()

        for flow in result.content["flows"]:
            assert flow["direction"] in ("input", "output")

    def test_amounts_extracted(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        _create_ldi_xlsx(xlsx)
        source = ExcelLCISource(xlsx, profile=SIMPLE_PROFILE)
        result = source.extract()

        water_flows = [f for f in result.content["flows"] if f["category"] == "Air"]
        assert len(water_flows) == 2
        amounts = {f["amount"] for f in water_flows}
        assert 1000.0 in amounts
        assert 500.0 in amounts

    def test_processes_tracked(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        _create_ldi_xlsx(xlsx)
        source = ExcelLCISource(xlsx, profile=SIMPLE_PROFILE)
        result = source.extract()

        processes = result.content["summary"]["processes"]
        assert "Well Operation" in processes
        assert "NSOP" in processes

    def test_summary_counts(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        _create_ldi_xlsx(xlsx)
        source = ExcelLCISource(xlsx, profile=SIMPLE_PROFILE)
        result = source.extract()

        assert result.content["summary"]["total_flows"] == 5
        assert result.metadata["flow_count"] == 5

    def test_original_category_preserved(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        _create_ldi_xlsx(xlsx)
        source = ExcelLCISource(xlsx, profile=SIMPLE_PROFILE)
        result = source.extract()

        water_flow = next(f for f in result.content["flows"] if f["category"] == "Air")
        assert water_flow["original_category"] == "Water"

    def test_helper_data_extracted(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        _create_ldi_xlsx_with_helpers(xlsx)
        source = ExcelLCISource(xlsx, profile=SIMPLE_PROFILE)
        result = source.extract()

        helper = result.content["helper_data"]
        assert "Projected Lifetime of Infrastructure" in helper
        assert helper["Projected Lifetime of Infrastructure"][0]["value"] == 20.0

    def test_excluded_categories_skipped(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        _create_ldi_xlsx_with_helpers(xlsx)
        source = ExcelLCISource(xlsx, profile=SIMPLE_PROFILE)
        result = source.extract()

        categories = {f["category"] for f in result.content["flows"]}
        assert "Raw Material from Processes" not in categories

    def test_empty_rows_skipped(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        _create_ldi_xlsx(
            xlsx,
            rows=[
                ["Well Op", "Water", "Produced", "Input", "L", 100.0],
                [None, None, None, None, None, None],
                ["NSOP", "Electricity", "Pompa", "Input", "kWh", 200.0],
            ],
        )
        source = ExcelLCISource(xlsx, profile=SIMPLE_PROFILE)
        result = source.extract()

        assert result.content["summary"]["total_flows"] == 2


# ---------------------------------------------------------------------------
# Tests: Extract Structure (no profile)
# ---------------------------------------------------------------------------


class TestExtractStructure:
    def test_returns_excel_profile(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        _create_ldi_xlsx(xlsx)
        source = ExcelLCISource(xlsx, profile=None)
        result = source.extract()

        assert result.content_type == "excel_profile"
        assert result.content["mode"] == "structure_extraction"
        assert "excel_profile" in result.content

    def test_profile_has_sheet_info(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        _create_ldi_xlsx(xlsx)
        source = ExcelLCISource(xlsx)
        result = source.extract()

        excel_profile = result.content["excel_profile"]
        assert "sheet_names" in excel_profile
        assert "sheets" in excel_profile
        assert "LDI Master" in excel_profile["sheet_names"]

    def test_metadata_has_sheet_count(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        _create_ldi_xlsx(xlsx)
        source = ExcelLCISource(xlsx)
        result = source.extract()

        assert result.metadata["sheet_count"] == 1


# ---------------------------------------------------------------------------
# Tests: Static helpers
# ---------------------------------------------------------------------------


class TestHelpers:
    def test_get_cell_valid(self):
        row = ["a", "b", "c"]
        assert ExcelLCISource._get_cell(row, {"index": 1}) == "b"

    def test_get_cell_none_spec(self):
        assert ExcelLCISource._get_cell(["a"], None) is None

    def test_get_cell_out_of_range(self):
        assert ExcelLCISource._get_cell(["a"], {"index": 5}) is None

    def test_get_numeric_valid(self):
        assert ExcelLCISource._get_numeric(["text", 42.5], {"index": 1}) == 42.5

    def test_get_numeric_string(self):
        assert ExcelLCISource._get_numeric(["abc"], {"index": 0}) is None

    def test_get_numeric_none(self):
        assert ExcelLCISource._get_numeric([None], {"index": 0}) is None

    def test_header_to_field_simple(self):
        assert ExcelLCISource._header_to_field("Data Source") == "data_source"

    def test_header_to_field_multi_word(self):
        assert ExcelLCISource._header_to_field("Material Composition") == "material_composition"

    def test_header_to_field_boolean(self):
        assert ExcelLCISource._header_to_field("Is Amount Balanced") == "is_amount_balanced"

    def test_header_to_field_short(self):
        assert ExcelLCISource._header_to_field("PIC") == "pic"

    def test_header_to_field_abbreviation(self):
        assert ExcelLCISource._header_to_field("TeR") == "te_r"


# ---------------------------------------------------------------------------
# Tests: Extra column extraction
# ---------------------------------------------------------------------------


class TestExtraColumnExtraction:
    """Test that unmapped Excel columns are preserved on flow dicts."""

    def _create_xlsx_with_extras(self, path: Path):
        """Create Excel with both mapped and unmapped columns."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "LDI Master"
        ws.append(
            [
                "Process",
                "Category",
                "Material",
                "Direction",
                "Unit",
                "Amount",
                "Data Source",
                "PIC",
                "Notes",
                "Review Status",
            ]
        )
        ws.append(
            [
                "Well Op",
                "Water",
                "PDAM",
                "Input",
                "L",
                1000.0,
                "Measured",
                "Budi",
                "Test note",
                "C",
            ]
        )
        ws.append(
            [
                "NSOP",
                "Electricity",
                "PLN",
                "Input",
                "kWh",
                500.0,
                "Estimated",
                "Andi",
                None,
                "P",
            ]
        )
        wb.save(str(path))

    def test_extra_columns_present(self, tmp_path):
        """Unmapped columns should appear as extra fields on flows."""
        xlsx = tmp_path / "test.xlsx"
        self._create_xlsx_with_extras(xlsx)
        source = ExcelLCISource(xlsx, profile=SIMPLE_PROFILE)
        result = source.extract()
        flows = result.content["flows"]

        water_flow = next(f for f in flows if f["flow_name"] == "PDAM")
        assert water_flow["data_source"] == "Measured"
        assert water_flow["pic"] == "Budi"
        assert water_flow["notes"] == "Test note"
        assert water_flow["review_status"] == "C"

    def test_none_extra_columns_skipped(self, tmp_path):
        """None values in extra columns should not be included."""
        xlsx = tmp_path / "test.xlsx"
        self._create_xlsx_with_extras(xlsx)
        source = ExcelLCISource(xlsx, profile=SIMPLE_PROFILE)
        result = source.extract()
        flows = result.content["flows"]

        elec_flow = next(f for f in flows if f["flow_name"] == "PLN")
        assert "notes" not in elec_flow  # None was skipped
        assert elec_flow["review_status"] == "P"

    def test_core_fields_unchanged(self, tmp_path):
        """Core mapped fields should work exactly as before."""
        xlsx = tmp_path / "test.xlsx"
        self._create_xlsx_with_extras(xlsx)
        source = ExcelLCISource(xlsx, profile=SIMPLE_PROFILE)
        result = source.extract()
        flows = result.content["flows"]

        water_flow = next(f for f in flows if f["flow_name"] == "PDAM")
        assert water_flow["category"] == "Air"
        assert water_flow["process"] == "Well Op"
        assert water_flow["direction"] == "input"
        assert water_flow["amount"] == 1000.0
        assert water_flow["unit"] == "L"

    def test_no_extra_columns(self, tmp_path):
        """Files with no extra columns should still parse correctly."""
        xlsx = tmp_path / "test.xlsx"
        _create_ldi_xlsx(xlsx)  # Only 6 columns, all mapped
        source = ExcelLCISource(xlsx, profile=SIMPLE_PROFILE)
        result = source.extract()
        assert len(result.content["flows"]) == 5

    def test_downstream_tools_unaffected(self, tmp_path):
        """Extra fields should pass through aggregate_flows without issues."""
        import json

        from lci_ignite.intelligence.tools import (
            _read_pipeline,
            aggregate_flows,
            clear_pipeline_data,
            clear_render_items,
        )

        clear_render_items()
        clear_pipeline_data()

        flows = [
            {
                "category": "Air",
                "flow_name": "PDAM",
                "amount": 100,
                "unit": "L",
                "process": "A",
                "data_source": "Measured",
                "pic": "Budi",
            },
            {
                "category": "Air",
                "flow_name": "PDAM",
                "amount": 200,
                "unit": "L",
                "process": "B",
                "data_source": "Estimated",
                "pic": "Andi",
            },
        ]
        aggregate_flows(json.dumps(flows))
        pipeline = _read_pipeline()
        merged = pipeline["flows"][0]

        # Core fields aggregated correctly
        assert merged["amount"] == 300
        assert merged["process"] == "A, B"
        # Extra fields preserved from first occurrence
        assert "data_source" in merged or "pic" in merged

        clear_render_items()
        clear_pipeline_data()
