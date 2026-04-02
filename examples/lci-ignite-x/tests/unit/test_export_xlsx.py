"""Tests for export_to_xlsx — PROPER template layout.

New layout (no Area/Process column):
    A: empty
    B(2): Input/Output
    C(3): Total (amount)
    D(4): Unit
    E+(5+): per product: Jumlah/FU | Unit | %  (3 cols per product)
"""

import openpyxl
import pytest

from lci_ignite.intelligence.tools import (
    _upload_dir_var,
    clear_pipeline_data,
    clear_render_items,
    export_to_xlsx,
    get_render_items,
)


@pytest.fixture(autouse=True)
def _clean(tmp_path):
    clear_render_items()
    clear_pipeline_data()
    _upload_dir_var.set(str(tmp_path))
    yield
    clear_render_items()
    clear_pipeline_data()


def _set_pipeline(flows, products=None):
    """Inject pipeline data for the tool."""
    from lci_ignite.intelligence.tools import _store_pipeline

    data = {"flows": flows}
    if products is not None:
        data["products"] = products
    _store_pipeline(data)


def _open_output(tmp_path, filename="io_table.xlsx"):
    wb = openpyxl.load_workbook(str(tmp_path / filename))
    return wb[wb.sheetnames[0]]


class TestExportToXlsxLayout:
    """Verify column layout — no Area column."""

    def test_column_a_empty(self, tmp_path):
        flows = [{"category": "Air", "flow_name": "Water", "amount": 100, "unit": "L"}]
        _set_pipeline(flows)
        export_to_xlsx(filename="io_table.xlsx")
        ws = _open_output(tmp_path)
        for row in range(1, ws.max_row + 1):
            assert ws.cell(row=row, column=1).value is None, f"Row {row} col A not empty"

    def test_data_starts_at_column_b(self, tmp_path):
        flows = [{"category": "Air", "flow_name": "Water", "amount": 100, "unit": "L"}]
        _set_pipeline(flows)
        export_to_xlsx(filename="io_table.xlsx")
        ws = _open_output(tmp_path)
        assert ws.cell(row=4, column=2).value == "Input/Output"

    def test_total_at_c_unit_at_d(self, tmp_path):
        flows = [{"category": "Air", "flow_name": "Water", "amount": 100, "unit": "L"}]
        _set_pipeline(flows)
        export_to_xlsx(filename="io_table.xlsx")
        ws = _open_output(tmp_path)
        assert ws.cell(row=4, column=3).value == "Total"
        assert ws.cell(row=4, column=4).value == "Unit"

    def test_no_area_column(self, tmp_path):
        """Area/Process column should not exist."""
        flows = [
            {"category": "Air", "flow_name": "Water", "amount": 100, "unit": "L", "process": "SPU"},
        ]
        _set_pipeline(flows)
        export_to_xlsx(filename="io_table.xlsx")
        ws = _open_output(tmp_path)
        # Check that no header cell contains "Area"
        for col in range(1, 15):
            val = ws.cell(row=4, column=col).value
            assert val != "Area", f"Found 'Area' header at column {col}"


class TestHeaderRows:
    """Verify 5-row header structure."""

    def test_row1_empty(self, tmp_path):
        _set_pipeline([{"category": "Air", "flow_name": "W", "amount": 1, "unit": "L"}])
        export_to_xlsx(filename="io_table.xlsx")
        ws = _open_output(tmp_path)
        for col in range(1, 10):
            assert ws.cell(row=1, column=col).value is None

    def test_row2_product_names(self, tmp_path):
        flows = [{"category": "Air", "flow_name": "W", "amount": 1, "unit": "L"}]
        products = [{"name": "Gas"}, {"name": "Oil"}]
        _set_pipeline(flows, products)
        export_to_xlsx(filename="io_table.xlsx")
        ws = _open_output(tmp_path)
        # Products start at col 5, 3 cols per product
        assert ws.cell(row=2, column=5).value == "Gas"
        assert ws.cell(row=2, column=8).value == "Oil"

    def test_row3_energy_values(self, tmp_path):
        flows = [{"category": "Air", "flow_name": "W", "amount": 1, "unit": "L"}]
        products = [
            {"name": "Gas", "total_energy_mj": 169594709007},
            {"name": "Oil", "total_energy_mj": 12792363977},
        ]
        _set_pipeline(flows, products)
        export_to_xlsx(filename="io_table.xlsx")
        ws = _open_output(tmp_path)
        assert ws.cell(row=3, column=4).value == "ALL PHM"
        assert ws.cell(row=3, column=5).value == 169594709007
        assert ws.cell(row=3, column=8).value == 12792363977

    def test_row5_subheaders(self, tmp_path):
        flows = [{"category": "Air", "flow_name": "W", "amount": 1, "unit": "L"}]
        products = [{"name": "Gas"}]
        _set_pipeline(flows, products)
        export_to_xlsx(filename="io_table.xlsx")
        ws = _open_output(tmp_path)
        assert ws.cell(row=5, column=3).value == "(sesuai periode kajian)"
        assert ws.cell(row=5, column=5).value == "Jumlah/FU"
        assert ws.cell(row=5, column=6).value == "Unit"
        assert ws.cell(row=5, column=7).value == "%"


class TestMergedCells:
    """Verify merged cells."""

    def test_header_vertical_merges(self, tmp_path):
        flows = [{"category": "Air", "flow_name": "W", "amount": 1, "unit": "L"}]
        products = [{"name": "Gas"}]
        _set_pipeline(flows, products)
        export_to_xlsx(filename="io_table.xlsx")
        ws = _open_output(tmp_path)
        merged = [str(m) for m in ws.merged_cells.ranges]
        assert "B4:B5" in merged
        assert "D4:D5" in merged

    def test_product_header_horizontal_merge(self, tmp_path):
        flows = [{"category": "Air", "flow_name": "W", "amount": 1, "unit": "L"}]
        products = [{"name": "Gas"}]
        _set_pipeline(flows, products)
        export_to_xlsx(filename="io_table.xlsx")
        ws = _open_output(tmp_path)
        merged = [str(m) for m in ws.merged_cells.ranges]
        # Product "Gas" at E4:G4 (3 columns)
        assert "E4:G4" in merged


class TestDataContent:
    """Verify data row content."""

    def test_flow_data_row(self, tmp_path):
        flows = [
            {
                "category": "Air",
                "flow_name": "Water",
                "amount": 1000,
                "unit": "L",
                "fu_per_mj_Gas": 0.004,
                "pct_Gas": 60.0,
            }
        ]
        products = [{"name": "Gas", "total_energy_mj": 100000}]
        _set_pipeline(flows, products)
        export_to_xlsx(filename="io_table.xlsx")
        ws = _open_output(tmp_path)
        # Data row at row 7 (row 6 = section header "Air")
        assert ws.cell(row=7, column=2).value == "Water"
        assert ws.cell(row=7, column=3).value == 1000
        assert ws.cell(row=7, column=4).value == "L"
        assert ws.cell(row=7, column=5).value == pytest.approx(0.004)
        assert ws.cell(row=7, column=6).value == "L/MJ"
        assert ws.cell(row=7, column=7).value == 60.0

    def test_total_row_pct_100(self, tmp_path):
        flows = [
            {
                "category": "Air",
                "flow_name": "W1",
                "amount": 100,
                "unit": "L",
                "fu_per_mj_Gas": 0.1,
                "pct_Gas": 60.0,
            },
            {
                "category": "Air",
                "flow_name": "W2",
                "amount": 200,
                "unit": "L",
                "fu_per_mj_Gas": 0.2,
                "pct_Gas": 40.0,
            },
        ]
        products = [{"name": "Gas"}]
        _set_pipeline(flows, products)
        export_to_xlsx(filename="io_table.xlsx")
        ws = _open_output(tmp_path)
        for row in range(6, ws.max_row + 1):
            if ws.cell(row=row, column=2).value == "Total":
                assert ws.cell(row=row, column=7).value == 1
                break


class TestEmissionSummary:
    """Verify Emisi Udara summary has no %."""

    def test_emission_summary_no_pct(self, tmp_path):
        flows = [
            {
                "category": "Emisi Udara",
                "flow_name": "CO2 Flaring",
                "amount": 1000,
                "unit": "kg",
                "fu_per_mj_Gas": 0.01,
                "pct_Gas": 80.0,
            },
            {
                "category": "Emisi Udara",
                "flow_name": "CO2 Engine",
                "amount": 500,
                "unit": "kg",
                "fu_per_mj_Gas": 0.005,
                "pct_Gas": 20.0,
            },
        ]
        products = [{"name": "Gas"}]
        _set_pipeline(flows, products)
        export_to_xlsx(filename="io_table.xlsx")
        ws = _open_output(tmp_path)
        emisi_header_row = None
        for row in range(6, ws.max_row + 1):
            if ws.cell(row=row, column=2).value == "Emisi Udara":
                emisi_header_row = row
                break
        assert emisi_header_row is not None
        data_row = emisi_header_row + 1
        if ws.cell(row=data_row, column=2).value:
            # % column (col 7) should be None for emission summary
            assert ws.cell(row=data_row, column=7).value is None


class TestSectionNaming:
    """Verify section naming matches PROPER template."""

    def test_sampah_section(self, tmp_path):
        flows = [
            {"category": "Sampah", "flow_name": "Sisa Pipa", "amount": 50, "unit": "kg"},
        ]
        _set_pipeline(flows)
        export_to_xlsx(filename="io_table.xlsx")
        ws = _open_output(tmp_path)
        found = False
        for row in range(6, ws.max_row + 1):
            if ws.cell(row=row, column=2).value == "Sampah":
                found = True
                break
        assert found, "Sampah section not found"


class TestFileOutput:
    """Verify file output and render item."""

    def test_returns_success_message(self, tmp_path):
        flows = [{"category": "Air", "flow_name": "W", "amount": 1, "unit": "L"}]
        _set_pipeline(flows)
        result = export_to_xlsx(filename="io_table.xlsx")
        assert "Excel IO Table exported" in result
        assert "io_table.xlsx" in result

    def test_file_created(self, tmp_path):
        flows = [{"category": "Air", "flow_name": "W", "amount": 1, "unit": "L"}]
        _set_pipeline(flows)
        export_to_xlsx(filename="io_table.xlsx")
        assert (tmp_path / "io_table.xlsx").exists()

    def test_render_item_pushed(self, tmp_path):
        flows = [{"category": "Air", "flow_name": "W", "amount": 1, "unit": "L"}]
        _set_pipeline(flows)
        export_to_xlsx(filename="io_table.xlsx")
        items = get_render_items()
        xlsx_items = [i for i in items if i.get("name") == "io_table.xlsx"]
        assert len(xlsx_items) == 1
        assert (
            xlsx_items[0]["mimeType"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


class TestFUUnitPerOutputUnit:
    """Verify FU unit columns use dynamic labels from fu_unit_labels."""

    def test_fu_unit_column_per_output_unit(self, tmp_path):
        """With per_output_unit mode, unit column should show /Barrel, not /MJ."""
        flows = [
            {
                "category": "Air",
                "flow_name": "Water",
                "amount": 1000,
                "unit": "L",
                "fu_per_mj_Minyak": 0.004,
                "pct_Minyak": 100.0,
            }
        ]
        products = [
            {
                "name": "Minyak",
                "total_energy_mj": 2212001236,
                "fu_unit_factor": 5992.74,
                "output_unit": "Barrel",
            }
        ]
        from lci_ignite.intelligence.tools import _store_pipeline

        pipeline_data = {
            "flows": flows,
            "products": products,
            "fu_mode": "per_output_unit",
            "fu_unit_labels": {"Minyak": "Barrel"},
        }
        _store_pipeline(pipeline_data)

        export_to_xlsx(filename="io_table.xlsx")
        ws = _open_output(tmp_path)
        # Data row: col 6 should be "L/Barrel" not "L/MJ"
        for row in range(6, ws.max_row + 1):
            if ws.cell(row=row, column=2).value == "Water":
                assert ws.cell(row=row, column=6).value == "L/Barrel"
                break
        else:
            pytest.fail("Water row not found")

    def test_row3_output_quantity_per_output_unit(self, tmp_path):
        """Row 3 should show total output quantity, not total_energy_mj."""
        flows = [{"category": "Air", "flow_name": "W", "amount": 1, "unit": "L"}]
        products = [
            {
                "name": "Gas",
                "total_energy_mj": 93776048,
                "fu_unit_factor": 1055055.85,
                "output_unit": "MMSCF",
            }
        ]
        from lci_ignite.intelligence.tools import _store_pipeline

        pipeline_data = {
            "flows": flows,
            "products": products,
            "fu_mode": "per_output_unit",
            "fu_unit_labels": {"Gas": "MMSCF"},
        }
        _store_pipeline(pipeline_data)

        export_to_xlsx(filename="io_table.xlsx")
        ws = _open_output(tmp_path)
        expected = 93776048 / 1055055.85
        assert ws.cell(row=3, column=5).value == pytest.approx(expected)
